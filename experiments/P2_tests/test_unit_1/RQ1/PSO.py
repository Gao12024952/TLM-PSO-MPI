from mpi4py import MPI
import random
import time
import copy
import math
import openpyxl
from openpyxl import Workbook
import os

# ==================== 实验配置 ====================
NUM_EXPERIMENTS = 20  # 每组运行的实验次数


# ================================================

def worker_loop(local_comm):
    rank = local_comm.Get_rank()
    status = MPI.Status()

    while True:
        # 使用 Probe 先检查消息标签
        local_comm.Probe(source=0, tag=MPI.ANY_TAG, status=status)
        tag = status.Get_tag()

        # 1. 优先处理控制消息
        if tag == 0:
            control_msg = local_comm.recv(source=0, tag=0, status=status)
            if control_msg == "QUIT":
                break
            continue

        # 2. 如果不是控制消息，则按顺序接收数据并直接赋给具体变量
        if rank == 1:
            # 接收: 温度(x), 湿度(y), 气压(z)
            x = local_comm.recv(source=0, tag=1, status=status)
            y = local_comm.recv(source=0, tag=2, status=status)
            z = local_comm.recv(source=0, tag=3, status=status)

            path = []
            if (x ** 2 / 100 > y * z / 20) != (x * 2 / 100 > y * z / 20):
                path.append(1)
            if (x ** 2 / 100 > y * z / 20) != (x ** 1.5 / 100 > y * z / 20):
                path.append(2)
            if (x ** 2 / 100 > y * z / 20) != (x ** 2 / 200 > y * z / 20):
                path.append(3)
            if (x ** 2 / 100 > y * z / 20) != (x ** 2 / 100 > y * z + 20):
                path.append(4)
            if (x ** 2 / 100 > y * z / 20) != (x ** 2 / 100 > y * z * 20):
                path.append(5)
            if (x ** 2 / 100 > y * z / 20) != (x ** 2 / 100 > y * z / 5):
                path.append(6)
            if (x ** 2 / 100 > y * z / 20) != (x ** 2 / 100 > y * y / 2):
                path.append(7)
            if (x ** 2 / 100 > y * z / 20) != (x ** 2 / 100 > y * x / 20):
                path.append(8)
            if (x ** 2 / 100 > y * z / 20) != (x * y / 100 > y * z / 20):
                path.append(9)
            if (x ** 2 / 100 > y * z / 20) != ((x - 100) ** 2 / 100 > y * z / 20):
                path.append(10)

            if x ** 2 / 100 > y * z / 20:
                type_code = 'B4'

            if ((x + y) / 2 > z / 3 + 150) != ((x + y - 280) / 2 > z / 3 + 150):
                path.append(11)
            if ((x + y) / 2 > z / 3 + 150) != ((x + y) / 2 > z + 3 + 150):
                path.append(12)
            if ((x + y) / 2 > z / 3 + 150) != ((x + y) / 2 > z + 150):
                path.append(13)
            if ((x + y) / 2 > z / 3 + 150) != ((x + y) / 2 > y * 1.2 + 150):
                path.append(14)
            if ((x + y) / 2 > z / 3 + 150) != ((x + y) / 2 > z / 3 + 380):
                path.append(15)
            if ((x + y) / 2 > z / 3 + 150) != ((x - y) / 2 > z / 3 + 150):
                path.append(16)
            if ((x + y) / 2 > z / 3 + 150) != ((x + y) / 2 > x / 4 + 150):
                path.append(17)
            if ((x + y) / 2 > z / 3 + 150) != ((x + x / 1.5) / 2 > z / 3 + 150):
                path.append(18)
            if ((x + y) / 2 > z / 3 + 150) != ((x + y) / 2 > 100 + 150):
                path.append(19)
            if ((x + y) / 2 > z / 3 + 150) != ((x + y) / 5 > z / 3 + 150):
                path.append(20)

            if (x + y) / 2 > z / 3 + 150:
                type_code = 'B5'

            if (x > (y - 50) * 4 and z > 600) != (x > (y - 50) + 4 and z > 600):
                path.append(21)
            if (x > (y - 50) * 4 and z > 600) != (x > (y - 50) / 4 and z > 600):
                path.append(22)
            if (x > (y - 50) * 4 and z > 600) != (x > (y - 50) * 8 and z > 600):
                path.append(23)
            if (x > (y - 50) * 4 and z > 600) != (x > (y * 1.5 - 50) * 4 and z > 600):
                path.append(24)
            if (x > (y - 50) * 4 and z > 600) != (x > (y - 50) * 4 and z > 450):
                path.append(25)
            if (x > (y - 50) * 4 and z > 600) != (x > (y - 10) * 4 and z > 600):
                path.append(26)
            if (x > (y - 50) * 4 and z > 600) != (x * 10 > (y - 50) * 4 and z > 600):
                path.append(27)
            if (x > (y - 50) * 4 and z > 600) != (x > (y - 50) * 4 and z + x / 2 > 600):
                path.append(28)
            if (x > (y - 50) * 4 and z > 600) != (x > (y - 50) * 4 and z + z / 5 > 600):
                path.append(29)
            if (x > (y - 50) * 4 and z > 600) != (x > (y - 50 + z / 25) * 4 and z > 600):
                path.append(30)

            if x > (y - 50) * 4 and z > 600:
                type_code = 'B3'

            local_comm.send(path, dest=0, tag=11)

        elif rank == 2:
            # 接收: 湿度(y), 气压(z), 风速(w)
            y = local_comm.recv(source=0, tag=1, status=status)
            z = local_comm.recv(source=0, tag=2, status=status)
            w = local_comm.recv(source=0, tag=3, status=status)

            path = []

            if (y * w / 100 > z / 8 + 50) != (y + w / 100 > z / 8 + 50):
                path.append(1)
            if (y * w / 100 > z / 8 + 50) != (y * w % 100 > z / 8 + 50):
                path.append(2)
            if (y * w / 100 > z / 8 + 50) != (y * w / 100 > z + 8 + 50):
                path.append(3)
            if (y * w / 100 > z / 8 + 50) != (y * w / 100 > z / 8 * 50):
                path.append(4)
            if (y * w / 100 > z / 8 + 50) != (y * w / 100 > z / 2 + 50):
                path.append(5)
            if (y * w / 100 > z / 8 + 50) != (y * y / 100 > z / 8 + 50):
                path.append(6)
            if (y * w / 100 > z / 8 + 50) != (w * w / 100 > z / 8 + 50):
                path.append(7)
            if (y * w / 100 > z / 8 + 50) != (y * w / 100 > y / 1.5 + 50):
                path.append(8)
            if (y * w / 100 > z / 8 + 50) != (y * w / 100 > w / 2 + 50):
                path.append(9)
            if (y * w / 100 > z / 8 + 50) != (y * w / 150 > z / 8 + 50):
                path.append(10)

            if y * w / 100 > z / 8 + 50:
                type_code = 'C4'

            if ((y + z / 10) % 20 < w / 10) != ((y + z / 9.8) % 20 < w / 10):
                path.append(11)
            if ((y + z / 10) % 20 < w / 10) != ((y + (z - 12) / 10) % 20 < w / 10):
                path.append(12)
            if ((y + z / 10) % 20 < w / 10) != ((y + z / 10 - 1.5) % 20 < w / 10):
                path.append(13)
            if ((y + z / 10) % 20 < w / 10) != ((y + z / 10) % 20 < w / 15):
                path.append(14)
            if ((y + z / 10) % 20 < w / 10) != ((y + z / 10) % 20 < (w - 28) / 10):
                path.append(15)
            if ((y + z / 10) % 20 < w / 10) != ((y + z / 10) % 20 < (y - 20) / 10):
                path.append(16)
            if ((y + z / 10) % 20 < w / 10) != ((y + z / 10) % 20 < z / 72):
                path.append(17)
            if ((y + z / 10) % 20 < w / 10) != ((y + z / 10) % 20 < 60 / 10):
                path.append(18)
            if ((y + z / 10) % 20 < w / 10) != ((y + z / 10 - w / 80) % 20 < w / 10):
                path.append(19)
            if ((y + z / 10) % 20 < w / 10) != ((y + z / 10 - y / 80) % 20 < w / 10):
                path.append(20)

            if (y + z / 10) % 20 < w / 10:
                type_code = 'C5'

            if (w ** 2 / 50 > y + z / 15) != (w ** 2 / 50 > y + z / 50):
                path.append(21)
            if (w ** 2 / 50 > y + z / 15) != (w ** 2 / 50 > y / 1.5 + z / 15):
                path.append(22)
            if (w ** 2 / 50 > y + z / 15) != (w ** 2 / 50 > y + z % 15):
                path.append(23)
            if (w ** 2 / 50 > y + z / 15) != (w ** 2 / 50 > z / 8 + z / 15):
                path.append(24)
            if (w ** 2 / 50 > y + z / 15) != (w ** 2 / 50 > w / 1.2 + z / 15):
                path.append(25)
            if (w ** 2 / 50 > y + z / 15) != (w ** 2 / 50 > y + y / 15):
                path.append(26)
            if (w ** 2 / 50 > y + z / 15) != (w ** 2 / 50 > y + w / 15):
                path.append(27)
            if (w ** 2 / 50 > y + z / 15) != (w * (w + 15) / 50 > y + z / 15):
                path.append(28)
            if (w ** 2 / 50 > y + z / 15) != (w * (y + 5) / 50 > y + z / 15):
                path.append(29)
            if (w ** 2 / 50 > y + z / 15) != (w ** 2 / 40 > y + z / 15):
                path.append(30)

            if w ** 2 / 50 > y + z / 15:
                type_code = 'C6'

            if (z / (w + 1) < y / 2 - 20) != (z / (w + 1) < y / 2.8 - 20):
                path.append(31)
            if (z / (w + 1) < y / 2 - 20) != (z / (w / 2 + 1) < y / 2 - 20):
                path.append(32)
            if (z / (w + 1) < y / 2 - 20) != (z * 2 / (w + 1) < y / 2 - 20):
                path.append(33)
            if (z / (w + 1) < y / 2 - 20) != (w * 18 / (w + 1) < y / 2 - 20):
                path.append(34)
            if (z / (w + 1) < y / 2 - 20) != (z / (w + 1) < (y - 20) / 2 - 20):
                path.append(35)
            if (z / (w + 1) < y / 2 - 20) != (z / (w + 1) < (z - 200) / 2 - 20):
                path.append(36)
            if (z / (w + 1) < y / 2 - 20) != (z / (w + 1) < w / 2.5 - 20):
                path.append(37)
            if (z / (w + 1) < y / 2 - 20) != ((z + 800) / (w + 1) < y / 2 - 20):
                path.append(38)
            if (z / (w + 1) < y / 2 - 20) != (z / (z / 18 + 1) < y / 2 - 20):
                path.append(39)
            if (z / (w + 1) < y / 2 - 20) != (z / (y / 2 + 1) < y / 2 - 20):
                path.append(40)

            if z / (w + 1) < y / 2 - 20:
                type_code = 'C7'

            local_comm.send(path, dest=0, tag=21)

        elif rank == 3:
            # 接收: 气压(z), 风速(w), 日照(m)
            z = local_comm.recv(source=0, tag=1, status=status)
            w = local_comm.recv(source=0, tag=2, status=status)
            m = local_comm.recv(source=0, tag=3, status=status)

            path = []
            if ((z + w * 3) / 2 < m ** 2 * 15) != ((z + w * 8) / 2 < m ** 2 * 15):
                path.append(1)
            if ((z + w * 3) / 2 < m ** 2 * 15) != ((z + w * 3) / 1.2 < m ** 2 * 15):
                path.append(2)
            if ((z + w * 3) / 2 < m ** 2 * 15) != ((z * 2 + w * 3) / 2 < m ** 2 * 15):
                path.append(3)
            if ((z + w * 3) / 2 < m ** 2 * 15) != ((z + w * 3) / 2 < m * 8 * 15):
                path.append(4)
            if ((z + w * 3) / 2 < m ** 2 * 15) != ((z + w * 3) / 2 < m ** 2 * 25):
                path.append(5)
            if ((z + w * 3) / 2 < m ** 2 * 15) != ((z + z * 1.2) / 2 < m ** 2 * 15):
                path.append(6)
            if ((z + w * 3) / 2 < m ** 2 * 15) != ((w + w * 3) / 2 < m ** 2 * 15):
                path.append(7)
            if ((z + w * 3) / 2 < m ** 2 * 15) != ((z + z + 100) / 2 < m ** 2 * 15):
                path.append(8)
            if ((z + w * 3) / 2 < m ** 2 * 15) != ((z + w * 3 + 500) / 2 < m ** 2 * 15):
                path.append(9)
            if ((z + w * 3) / 2 < m ** 2 * 15) != ((50 * m + w * 3) / 2 < m ** 2 * 15):
                path.append(10)

            if (z + w * 3) / 2 < m ** 2 * 15:
                type_code = 'D4'

            if (z % (m + 1) == 0 and w > 85) != (z + (m + 1) == 0 and w > 85):
                path.append(11)
            if (z % (m + 1) == 0 and w > 85) != (z * (m + 1) == 0 and w > 85):
                path.append(12)
            if (z % (m + 1) == 0 and w > 85) != (z % (m + 5) == 0 and w > 85):
                path.append(13)
            if (z % (m + 1) == 0 and w > 85) != (z % (m + z) == 0 and w > 85):
                path.append(14)
            if (z % (m + 1) == 0 and w > 85) != (z % (m + m) == 0 and w > 85):
                path.append(15)
            if (z % (m + 1) == 0 and w > 85) != (z % (m + w) == 0 and w > 85):
                path.append(16)
            if (z % (m + 1) == 0 and w > 85) != (m % (m + 1) == 0 and w > 85):
                path.append(17)
            if (z % (m + 1) == 0 and w > 85) != (w % (m + 1) == 0 and w > 85):
                path.append(18)
            if (z % (m + 1) == 0 and w > 85) != (z % (w + 1) == 0 and w > 85):
                path.append(19)
            if (z % (m + 1) == 0 and w > 85) != (z % (m + 1) == 0 and w / 2 > 85):
                path.append(20)
            if z % (m + 1) == 0 and w > 85:
                type_code = 'D3'

            local_comm.send(path, dest=0, tag=31)


# PSO 类
class PSO:
    # --- PSO 算法参数 ---
    DIMENSION = 5  # 维度：[温度, 湿度, 气压, 风速, 日照]
    SWARM_SIZE = 20  # 粒子群大小
    MAX_ITERATIONS = 3000  # 最大迭代次数
    MIN_SPEED = -100  # 最小速度
    MAX_SPEED = 100  # 最大速度
    W = 0.729  # 惯性权重

    # 为每个维度定义位置范围
    POSITION_RANGES = [
        (50, 500),  # 生产速度 (件/小时)
        (70, 100),  # 质量指标 (%)
        (100, 1000),  # 能耗 (kW)
        (60, 120),  # 人员效率 (%)
        (1, 10)  # 设备状态评分 (1-10分)
    ]

    def __init__(self, local_comm, group_color):
        self.rand = random.Random()
        self.comm = local_comm
        self.group_color = group_color
        self.swarm = [self.Particle() for _ in range(PSO.SWARM_SIZE)]
        self.gBest = self.Particle()
        self.gBestFitness = float('-inf')

        self.evaluation_count = 0

    class Particle:
        def __init__(self):
            self.position = [0] * PSO.DIMENSION
            self.fitness = float('-inf')
            self.pBest = [0] * PSO.DIMENSION
            self.pBestFitness = float('-inf')
            self.velocity = [0] * PSO.DIMENSION
            self.paths = [[] for _ in range(4)]

    def initializeSwarm(self):
        """初始化粒子群，使其在指定范围内生成"""
        samplesPerDimension = self.SWARM_SIZE
        for dim in range(self.DIMENSION):
            rangeMin, rangeMax = self.POSITION_RANGES[dim]

            samples = []
            intervalSize = (rangeMax - rangeMin) / samplesPerDimension
            for i in range(samplesPerDimension):
                start = rangeMin + i * intervalSize
                sample = start + self.rand.random() * intervalSize
                samples.append(sample)

            self.rand.shuffle(samples)
            for i in range(self.SWARM_SIZE):
                self.swarm[i].position[dim] = samples[i]
                self.swarm[i].velocity[dim] = self.MIN_SPEED + self.rand.random() * (self.MAX_SPEED - self.MIN_SPEED)

    def updateParticles(self, currentTargetPaths):
        """更新粒子群中所有粒子的速度、位置、适应度以及pBest和gBest"""
        for particle in self.swarm:
            # 1. 更新速度和位置
            for i in range(PSO.DIMENSION):
                r1 = self.rand.random()
                r2 = self.rand.random()
                particle.velocity[i] = self.W * particle.velocity[i] + \
                                       2.0 * r1 * (particle.pBest[i] - particle.position[i]) + \
                                       2.0 * r2 * (self.gBest.position[i] - particle.position[i])
                particle.velocity[i] = max(PSO.MIN_SPEED, min(PSO.MAX_SPEED, particle.velocity[i]))
                particle.position[i] += particle.velocity[i]

                rangeMin, rangeMax = self.POSITION_RANGES[i]
                particle.position[i] = max(rangeMin, min(rangeMax, particle.position[i]))

            # 2. 计算新位置的适应度
            mpi_result = self.generateMPIPaths(particle.position)
            particle.paths = mpi_result
            fitness = self.calculateFitnessBasedOnPaths(mpi_result, currentTargetPaths)
            particle.fitness = fitness

            # 3. 更新个体最优 (pBest)
            if fitness > particle.pBestFitness:
                particle.pBest = particle.position.copy()
                particle.pBestFitness = fitness

            # 4. 更新全局最优 (gBest)
            if fitness > self.gBestFitness:
                self.gBest = copy.deepcopy(particle)
                self.gBestFitness = fitness

    def generateMPIPaths(self, a):
        """主进程与工作进程通信，处理5个变量"""
        comm = self.comm
        status = MPI.Status()
        path = []
        x, y, z, w, m = a[0], a[1], a[2], a[3], a[4]

        comm.send(x, dest=1, tag=1)
        comm.send(y, dest=1, tag=2)
        comm.send(z, dest=1, tag=3)

        comm.send(y, dest=2, tag=1)
        comm.send(z, dest=2, tag=2)
        comm.send(w, dest=2, tag=3)

        comm.send(z, dest=3, tag=1)
        comm.send(w, dest=3, tag=2)
        comm.send(m, dest=3, tag=3)

        # 主进程的路径计算
        if (x / (m + 1) > y + z / 50) != (x % (m + 1) > y + z / 50):
            path.append(1)
        if (x / (m + 1) > y + z / 50) != (x / (m * 1) > y + z / 50):
            path.append(2)
        if (x / (m + 1) > y + z / 50) != (x / (m + 1) > y * z / 50):
            path.append(3)
        if (x / (m + 1) > y + z / 50) != (x / (m + 1) > y + z + 50):
            path.append(4)
        if (x / (m + 1) > y + z / 50) != (x / (m + 1) > y + z * 50):
            path.append(5)
        if (x / (m + 1) > y + z / 50) != (x / (m + 1) > z + z / 50):
            path.append(6)
        if (x / (m + 1) > y + z / 50) != (x / (m + 1) > x + z / 50):
            path.append(7)
        if (x / (m + 1) > y + z / 50) != (z / 5 / (m + 1) > y + z / 50):
            path.append(8)
        if (x / (m + 1) > y + z / 50) != (x / (m + 1) > y - z / 50):
            path.append(9)
        if (x / (m + 1) > y + z / 50) != (x / (m + 1) > y + x):
            path.append(10)

        if x / (m + 1) > y + z / 50:
            type_code = 'A4'

        if ((w + m * 8) ** 2 < x * y) != ((w + m * 5) ** 2 < x * y):
            path.append(11)
        if ((w + m * 8) ** 2 < x * y) != ((w + m * 8) * 100 < x * y):
            path.append(12)
        if ((w + m * 8) ** 2 < x * y) != ((w + m * 8) ** 2 < x * x / 2):
            path.append(13)
        if ((w + m * 8) ** 2 < x * y) != ((w + w / 2) ** 2 < x * y):
            path.append(14)
        if ((w + m * 8) ** 2 < x * y) != ((w + x / 6) ** 2 < x * y):
            path.append(15)
        if ((w + m * 8) ** 2 < x * y) != ((w + 5 * 8) ** 2 < x * y):
            path.append(16)
        if ((w + m * 8) ** 2 < x * y) != ((w + (m - 3) * 8) ** 2 < x * y):
            path.append(17)
        if ((w + m * 8) ** 2 < x * y) != ((w - 20 + m * 8) ** 2 < x * y):
            path.append(18)
        if ((w + m * 8) ** 2 < x * y) != ((w + m * 8) ** 2 < (x + 50) * y):
            path.append(19)
        if ((w + m * 8) ** 2 < x * y) != ((w + m * 8) ** 2 < x * 1.2 * y):
            path.append(20)

        if (w + m * 8) ** 2 < x * y:
            type_code = 'A5'

        if (z / 5 + x > y * 2 + w) != (z / 2 + x > y * 2 + w):
            path.append(21)
        if (z / 5 + x > y * 2 + w) != (x + x > y * 2 + w):
            path.append(22)
        if (z / 5 + x > y * 2 + w) != (z / 5 + x > y + 2 + w):
            path.append(23)
        if (z / 5 + x > y * 2 + w) != (z / 5 + x > y / 2 + w):
            path.append(24)
        if (z / 5 + x > y * 2 + w) != (y + x > y * 2 + w):
            path.append(25)
        if (z / 5 + x > y * 2 + w) != (w + x > y * 2 + w):
            path.append(26)
        if (z / 5 + x > y * 2 + w) != (z / 5 + x + 50 > y * 2 + w):
            path.append(27)
        if (z / 5 + x > y * 2 + w) != (z / 5 + x > y * 2.8 + w):
            path.append(28)
        if (z / 5 + x > y * 2 + w) != (z / 5 + x + z / 5 > y * 2 + w):
            path.append(29)
        if (z / 5 + x > y * 2 + w) != (z / 5 + x > y * 2 + w / 2):
            path.append(30)

        if z / 5 + x > y * 2 + w:
            type_code = 'A6'

        path1 = comm.recv(source=1, tag=11, status=status)
        path2 = comm.recv(source=2, tag=21, status=status)
        path3 = comm.recv(source=3, tag=31, status=status)

        return [path, path1, path2, path3]

    def calculateFitnessBasedOnPaths(self, mpi_paths, targetPaths):
        """计算生成的路径与目标路径的匹配度作为适应度"""
        self.evaluation_count += 1

        total_fitness = 0.0
        for i in range(4):
            target = targetPaths[i]
            if target:
                match_count = sum(1 for branch in target if branch in mpi_paths[i])
                local_fitness = match_count / len(target)
            else:
                local_fitness = 1.0 if not mpi_paths[i] else 0.0
            total_fitness += local_fitness
        return total_fitness / 4.0

    def run(self, currentTargetPaths):
        """执行PSO算法的主循环，返回结果"""
        start_time = MPI.Wtime()
        found = False
        actual_iterations = 0

        for iteration in range(PSO.MAX_ITERATIONS):
            self.updateParticles(currentTargetPaths)
            actual_iterations = iteration + 1

            if self.gBestFitness == 1.0:
                found = True
                break

        end_time = MPI.Wtime()
        elapsed = end_time - start_time
        return found, elapsed, actual_iterations, self.evaluation_count


def save_all_results_to_excel(all_group_results):
    """将所有组的实验结果保存到Excel文件中"""
    output_dir = "experimental_results"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    num_experiments = len(all_group_results[0]['success'])
    num_groups = len(all_group_results)

    file_types = [
        ("success", "是否成功"),
        ("iterations", "迭代次数"),
        ("evaluations", "评价次数"),
        ("time", "迭代时间"),
        ("fitness", "最终适应度")
    ]

    for file_type, description in file_types:
        wb = Workbook()
        ws = wb.active
        ws.title = description

        ws.cell(row=1, column=1, value="实验次数")
        for group_id in range(num_groups):
            ws.cell(row=1, column=group_id + 2, value=f"路径{group_id + 1}")

        for exp_id in range(num_experiments):
            ws.cell(row=exp_id + 2, column=1, value=exp_id + 1)
            for group_id in range(num_groups):
                value = all_group_results[group_id][file_type][exp_id]
                ws.cell(row=exp_id + 2, column=group_id + 2, value=value)

        filename = f"{output_dir}/{file_type}.xlsx"
        wb.save(filename)
        print(f"[全局] 已保存 {description} 数据到: {filename}")


def print_statistics(all_group_results, num_groups):
    """打印统计结果"""
    print("\n" + "=" * 80)
    print("实验统计结果".center(80))
    print("=" * 80)

    print("\n" + "-" * 80)
    print("各组统计".center(80))
    print("-" * 80)
    print(f"{'组ID':<8} {'成功率':<12} {'平均迭代':<12} {'平均评价':<12} {'平均时间':<12} {'平均适应度':<12}")
    print("-" * 80)

    group_stats = []
    for group_id in range(num_groups):
        results = all_group_results[group_id]
        success_count = sum(results['success'])
        total_count = len(results['success'])

        avg_iterations = sum(results['iterations']) / total_count
        avg_evaluations = sum(results['evaluations']) / total_count
        avg_time = sum(results['time']) / total_count
        avg_fitness = sum(results['fitness']) / total_count

        success_rate = success_count / total_count

        print(
            f"组{group_id:<6} {success_count}/{total_count:<9} {avg_iterations:<12.1f} {avg_evaluations:<12.0f} {avg_time:<12.4f}s {avg_fitness:<12.6f}")

        group_stats.append({
            'group_id': group_id,
            'success_rate': success_rate,
            'success_count': success_count,
            'total_count': total_count,
            'avg_iterations': avg_iterations,
            'avg_evaluations': avg_evaluations,
            'avg_time': avg_time,
            'avg_fitness': avg_fitness
        })

    print("\n" + "=" * 80)
    print("全局汇总".center(80))
    print("=" * 80)

    all_success = []
    all_iterations = []
    all_evaluations = []
    all_times = []
    all_fitness = []

    for group_results in all_group_results:
        all_success.extend(group_results['success'])
        all_iterations.extend(group_results['iterations'])
        all_evaluations.extend(group_results['evaluations'])
        all_times.extend(group_results['time'])
        all_fitness.extend(group_results['fitness'])

    total_experiments = len(all_success)
    total_success = sum(all_success)
    total_fail = total_experiments - total_success

    print(f"\n总实验次数: {total_experiments} ({num_groups}组 × {NUM_EXPERIMENTS}次)")
    print(f"总成功次数: {total_success} ({total_success / total_experiments * 100:.2f}%)")
    print(f"总失败次数: {total_fail} ({total_fail / total_experiments * 100:.2f}%)")

    print(f"\n迭代次数统计:")
    print(f"  平均: {sum(all_iterations) / len(all_iterations):.1f}")
    print(f"  中位数: {sorted(all_iterations)[len(all_iterations) // 2]:.1f}")
    print(f"  最小: {min(all_iterations)}")
    print(f"  最大: {max(all_iterations)}")

    print(f"\n评价次数统计:")
    print(f"  平均: {sum(all_evaluations) / len(all_evaluations):.0f}")
    print(f"  中位数: {sorted(all_evaluations)[len(all_evaluations) // 2]:.0f}")
    print(f"  最小: {min(all_evaluations)}")
    print(f"  最大: {max(all_evaluations)}")

    print(f"\n运行时间统计:")
    print(f"  平均: {sum(all_times) / len(all_times):.4f}s")
    print(f"  中位数: {sorted(all_times)[len(all_times) // 2]:.4f}s")
    print(f"  最小: {min(all_times):.4f}s")
    print(f"  最大: {max(all_times):.4f}s")

    print(f"\n适应度统计:")
    print(f"  平均: {sum(all_fitness) / len(all_fitness):.6f}")
    print(f"  中位数: {sorted(all_fitness)[len(all_fitness) // 2]:.6f}")
    print(f"  最小: {min(all_fitness):.6f}")
    print(f"  最大: {max(all_fitness):.6f}")

    sorted_by_success = sorted(group_stats, key=lambda x: x['success_rate'])
    print(
        f"\n最难优化的组: 组{sorted_by_success[0]['group_id']} (成功率 {sorted_by_success[0]['success_rate'] * 100:.1f}%)")
    print(
        f"最易优化的组: 组{sorted_by_success[-1]['group_id']} (成功率 {sorted_by_success[-1]['success_rate'] * 100:.1f}%)")

    print("\n" + "=" * 80)


def main():
    global_start = MPI.Wtime()

    targetPathsSets = [
        [[1, 3, 4, 5, 6, 7, 8, 10, 15], [1, 2, 4, 5, 6, 7, 9, 11, 12, 13, 14, 15, 16, 17, 19, 20],
         [2, 3, 4, 5, 6, 8, 9, 10, 14, 15, 16, 17, 18, 36], [4, 11, 12, 13, 14, 15, 16, 17, 19, 20]

         ],
        [[1, 3, 4, 5, 6, 7, 8, 10], [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 19, 20],
         [3, 4, 5, 6, 8, 9, 10, 11, 14, 15, 16, 17, 18, 36], [11, 12, 14, 16, 17, 19, 20]

         ],

        [[1, 3, 4, 5, 6, 7, 8, 10, 15], [1, 2, 4, 5, 6, 9, 11, 12, 13, 15, 20],
         [1, 3, 4, 5, 6, 8, 9, 10, 14, 16, 17, 18, 34, 36, 39], [1, 2, 9, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20]

         ],

        [[1, 3, 4, 5, 6, 7, 8, 10], [1, 2, 4, 5, 6, 7, 9, 11, 12, 13, 15, 16, 20],
         [2, 3, 4, 5, 8, 9, 10, 14, 15, 16, 17, 18, 36], [1, 2, 3, 9, 10, 11, 12, 13, 14, 16, 17, 18, 19, 20]

         ],

        [[1, 3, 4, 5, 6, 7, 8, 10, 28], [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 19, 20],
         [2, 3, 4, 5, 8, 9, 10, 36], [4, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20]

         ],
        [[1, 3, 4, 5, 6, 7, 8, 10, 14, 15, 16], [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 18],
         [1, 3, 4, 5, 6, 8, 9, 10, 14, 15, 16, 17, 18, 34, 36, 39], [11, 12, 13, 14, 15, 16, 17, 19, 20]

         ],

        [[1, 3, 4, 5, 6, 7, 8, 10], [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 19, 20],
         [3, 4, 5, 6, 8, 9, 10, 12, 13, 19, 20], [4, 11, 12, 13, 14, 16, 17, 19, 20]

         ],
        [[11, 12, 13, 16, 17, 18, 19, 20], [1, 2, 4, 5, 6, 7, 9, 11, 12, 13, 14, 15, 16, 17, 19, 20],
         [3, 4, 5, 6, 8, 9, 10, 17, 34, 36, 39], [11, 12, 13, 14, 15, 16, 17, 18, 19, 20]

         ],
        [[1, 3, 4, 5, 6, 7, 8, 10], [1, 2, 4, 5, 6, 7, 9, 11, 12, 13, 15, 16, 20],
         [1, 2, 3, 4, 5, 6, 8, 9, 10, 14, 15, 16, 17, 18], [4, 5, 7, 11, 12, 13, 14, 15, 16, 17, 19, 20]

         ],
        [[2], [1, 2, 3, 4, 5, 6, 7, 8, 9, 11, 12, 13, 14, 15, 16, 17, 19, 20],
         [2, 3, 4, 5, 8, 9, 10, 14, 15, 16, 17, 18, 36], [4, 5, 6, 7, 8, 11, 12, 14, 15, 16, 17, 18, 19, 20]

         ],
        [[2, 9], [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 19, 20],
         [2, 3, 4, 5, 6, 8, 9, 10, 14, 15, 16, 17, 18, 36], [4, 5, 7, 11, 12, 14, 15, 16, 17, 19, 20]

         ],

        [[22, 23, 24, 25, 26, 27, 30], [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 18], [2, 3, 4, 8, 11, 14, 15, 16, 17, 18, 36],
         [1, 2, 3, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20]

         ],

        [[1, 3, 4, 5, 6, 7, 8, 10], [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 19, 20],
         [3, 4, 5, 7, 8, 9, 10, 21, 22, 24, 25, 26, 27, 28, 29, 30, 36], [4]

         ],

        [[1, 3, 4, 5, 6, 7, 8, 10, 14, 15, 16, 21, 22, 23, 24, 25, 26, 27, 30], [1, 2, 3, 4, 5, 6, 7, 8, 9, 10],
         [2, 3, 4, 5, 8, 9, 36], [11, 12, 13, 14, 16, 17, 19, 20]

         ],
        [[1, 3, 4, 5, 6, 7, 8, 10, 15], [1, 2, 3, 4, 5, 6, 7, 9, 10, 14, 17, 18, 19, 25, 28, 29],
         [7, 16, 18, 34, 39, 40], [4, 7, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20]

         ],

        [[1, 3, 4, 5, 6, 7, 8, 10], [1, 2, 3, 4, 5, 6, 7, 9, 11, 12, 13, 15, 16, 17, 20],
         [14, 15, 17, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30], [1, 2, 3, 6, 8, 9]

         ],
        [[2, 9], [1, 2, 3, 4, 5, 6, 9, 10, 14, 17, 18, 19, 25, 28, 29],
         [15, 21, 22, 23, 25, 26, 27, 28, 29, 30, 31, 33, 34, 35, 38, 39, 40], [1, 2, 3, 6, 8, 9]

         ],
        [[2, 8, 9, 11, 17, 18, 19, 20, 25, 26], [23, 24, 26, 30],
         [12, 13, 17, 19, 20, 24, 29, 31, 32, 33, 34, 35, 38, 39, 40], [11, 12, 14, 16, 17, 19, 20]

         ],
        [[11, 12, 15, 17, 18, 19, 20, 25, 26, 28], [21, 22, 27], [21, 22, 23, 25, 26, 27, 28, 29, 30],
         [7, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20]

         ],

        [[11, 12, 13, 14, 15, 16, 17, 18, 19, 20], [25, 28, 29],
         [21, 22, 23, 26, 27, 28, 30, 31, 32, 33, 34, 35, 37, 38, 39, 40], [1, 2, 3, 6, 8, 9]

         ],
    ]

    comm_world = MPI.COMM_WORLD
    size = comm_world.Get_size()
    rank_world = comm_world.Get_rank()

    # 串行执行模式：只需要 4 个进程
    if size != 4:
        if rank_world == 0:
            print(f"错误：串行模式需要且仅需 4 个 MPI 进程（1主3从），当前进程数为 {size}。")
        MPI.Finalize()
        exit()

    local_comm = comm_world
    local_rank = rank_world
    num_paths = len(targetPathsSets)

    if rank_world == 0:
        print("\n" + "=" * 80)
        print("串行实验配置".center(80))
        print("=" * 80)
        print(f"进程总数: {size} (1个主进程 + 3个工作进程)")
        print(f"总路径集数: {num_paths}")
        print(f"每组实验次数: {NUM_EXPERIMENTS}")
        print(f"PSO参数: 粒子数={PSO.SWARM_SIZE}, 最大迭代={PSO.MAX_ITERATIONS}")
        print("=" * 80)
        print(f"\n[运行中] 正在依次执行 {num_paths} 组路径实验...\n")

        all_results = []

        # 外层循环：逐个路径集进行实验
        for set_idx in range(num_paths):
            currentTargetPaths = targetPathsSets[set_idx]
            print(f"[进度] 正在执行路径集 {set_idx + 1}/{num_paths}...")

            success_list = []
            iteration_list = []
            evaluations_list = []
            time_list = []
            fitness_list = []

            for exp_id in range(NUM_EXPERIMENTS):
                # 保持随机种子与原逻辑一致，但通过 set_idx 区分
                random.seed(set_idx * 10000 + exp_id)

                psoInstance = PSO(local_comm, group_color=set_idx)
                psoInstance.initializeSwarm()
                found, elapsed, actual_iterations, evaluations = psoInstance.run(currentTargetPaths)

                success_list.append(1 if found else 0)
                iteration_list.append(actual_iterations)
                evaluations_list.append(evaluations)
                time_list.append(elapsed)
                fitness_list.append(psoInstance.gBestFitness)

            # 存储该路径集的所有结果
            all_results.append({
                'success': success_list,
                'iterations': iteration_list,
                'evaluations': evaluations_list,
                'time': time_list,
                'fitness': fitness_list
            })

        # 所有路径循环结束后，通知 worker 进程退出
        for dest in range(1, 4):
            local_comm.send("QUIT", dest=dest, tag=0)

        # 打印统计并保存
        print_statistics(all_results, num_paths)
        save_all_results_to_excel(all_results)

        global_end = MPI.Wtime()
        print(f"\n程序总运行时间：{global_end - global_start:.4f} 秒\n")

    else:
        # 工作进程逻辑：持续待命，直到主进程发来 QUIT
        worker_loop(local_comm)

    MPI.Finalize()


if __name__ == "__main__":
    main()