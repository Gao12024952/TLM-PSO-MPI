from mpi4py import MPI
import random
import time
import copy
import math
import openpyxl
from openpyxl import Workbook
import os

# ==================== 实验配置 ====================
NUM_EXPERIMENTS = 20  # 每组运行的实验次数


# ================================================

def worker_loop(local_comm):
    rank = local_comm.Get_rank()
    status = MPI.Status()

    while True:
        # 使用 Probe 先检查消息标签
        local_comm.Probe(source=0, tag=MPI.ANY_TAG, status=status)
        tag = status.Get_tag()

        # 1. 优先处理控制消息
        if tag == 0:
            control_msg = local_comm.recv(source=0, tag=0, status=status)
            if control_msg == "QUIT":
                break
            continue

        # 2. 如果不是控制消息，则按顺序接收数据并直接赋给具体变量
        if rank == 1:
            # 接收: 温度(x), 湿度(y), 气压(z)
            x = local_comm.recv(source=0, tag=1, status=status)
            y = local_comm.recv(source=0, tag=2, status=status)
            z = local_comm.recv(source=0, tag=3, status=status)

            path = []
            if (x ** 2 + y ** 2 + z ** 2 > 5000) != (x * (x + 80) + y ** 2 + z ** 2 > 5000):
                path.append(1)
            if (x ** 2 + y ** 2 + z ** 2 > 5000) != (x ** 2 + y * (y + 50) + z ** 2 > 5000):
                path.append(2)
            if (x ** 2 + y ** 2 + z ** 2 > 5000) != (x ** 2 + y ** 2 + z * (z + 60) > 5000):
                path.append(3)
            if (x ** 2 + y ** 2 + z ** 2 > 5000) != (x ** 2 + y ** 2 + z ** 2 > 3000):
                path.append(4)
            if (x ** 2 + y ** 2 + z ** 2 > 5000) != (x ** 2.5 + y ** 2 + z ** 2 > 5000):
                path.append(5)
            if (x ** 2 + y ** 2 + z ** 2 > 5000) != (x ** 2 + y ** 2.8 + z ** 2 > 5000):
                path.append(6)
            if (x ** 2 + y ** 2 + z ** 2 > 5000) != (x * (x + y * 4) + y ** 2 + z ** 2 > 5000):
                path.append(7)
            if (x ** 2 + y ** 2 + z ** 2 > 5000) != (x ** 2 + y * (x + 20) + z ** 2 > 5000):
                path.append(8)
            if (x ** 2 + y ** 2 + z ** 2 > 5000) != (x ** 2 + y ** 2 + 58 ** 2 > 5000):
                path.append(9)
            if (x ** 2 + y ** 2 + z ** 2 > 5000) != (x ** 2 + y ** 2 + z ** 2 > 5000 - x * 80):
                path.append(10)

            if x ** 2 + y ** 2 + z ** 2 > 5000:
                type_code = 'B1'

            if ((x - y / 10) ** 3 > z * 200) != ((x * 1.5 - y / 10) ** 3 > z * 200):
                path.append(11)
            if ((x - y / 10) ** 3 > z * 200) != ((x - y / 5) ** 3 > z * 200):
                path.append(12)
            if ((x - y / 10) ** 3 > z * 200) != ((x - y / 10 + 8) ** 3 > z * 200):
                path.append(13)
            if ((x - y / 10) ** 3 > z * 200) != ((x - y / 10 + 8) ** 3 > z * 200):
                path.append(14)
            if ((x - y / 10) ** 3 > z * 200) != ((x - y / 10) ** 3 > z * 85):
                path.append(15)
            if ((x - y / 10) ** 3 > z * 200) != ((x - y / 10) ** 3 > (x - 2) * 200):
                path.append(16)
            if ((x - y / 10) ** 3 > z * 200) != ((x - 15 / 10) ** 3 > z * 200):
                path.append(17)
            if ((x - y / 10) ** 3 > z * 200) != ((x - y / 10) ** 3 > 20 * 200):
                path.append(18)
            if ((x - y / 10) ** 3 > z * 200) != ((x - y / 10) ** 3 > (z - 20) * 200):
                path.append(19)
            if ((x - y / 10) ** 3 > z * 200) != ((x - y / 10) ** 2 * 50 > z * 200):
                path.append(20)

            if (x - y / 10) ** 3 > z * 200:
                type_code = 'B2'

            if (y % x < z / (y / 10 + 1)) != (y % x < z * 1.5 / (y / 10 + 1)):
                path.append(21)
            if (y % x < z / (y / 10 + 1)) != (y % x < (z + 40) / (y / 10 + 1)):
                path.append(22)
            if (y % x < z / (y / 10 + 1)) != (y % x < z / (y / 5 + 1)):
                path.append(23)
            if (y % x < z / (y / 10 + 1)) != (y % x < z / (50 / 10 + 1)):
                path.append(24)
            if (y % x < z / (y / 10 + 1)) != (y % x < 80 / (y / 10 + 1)):
                path.append(25)
            if (y % x < z / (y / 10 + 1)) != (y % x < z / (x + 1)):
                path.append(26)
            if (y % x < z / (y / 10 + 1)) != (y % x < z / (y / 10 + 1) + 5):
                path.append(27)
            if (y % x < z / (y / 10 + 1)) != (y % x < z / (z / 3.5 + 1)):
                path.append(28)
            if (y % x < z / (y / 10 + 1)) != (y % x < (y / 2 + 10) / (y / 10 + 1)):
                path.append(29)
            if (y % x < z / (y / 10 + 1)) != (y % x < (x + 20) / (y / 10 + 1)):
                path.append(30)

            if y % x < z / (y / 10 + 1):
                type_code = 'B3'

            local_comm.send(path, dest=0, tag=11)

        elif rank == 2:
            # 接收: 湿度(y), 气压(z), 风速(w)
            y = local_comm.recv(source=0, tag=1, status=status)
            z = local_comm.recv(source=0, tag=2, status=status)
            w = local_comm.recv(source=0, tag=3, status=status)

            path = []

            if (w / (y + z) > (w / 20) ** 2) != (w * 35 / (y + z) > (w / 20) ** 2):
                path.append(1)
            if (w / (y + z) > (w / 20) ** 2) != ((w + 6000) / (y + z) > (w / 20) ** 2):
                path.append(2)
            if (w / (y + z) > (w / 20) ** 2) != (w / (y + z) + 100 > (w / 20) ** 2):
                path.append(3)
            if (w / (y + z) > (w / 20) ** 2) != (8000 / (y + z) > (w / 20) ** 2):
                path.append(4)
            if (w / (y + z) > (w / 20) ** 2) != (y * 100 / (y + z) > (w / 20) ** 2):
                path.append(5)
            if (w / (y + z) > (w / 20) ** 2) != (z * 150 / (y + z) > (w / 20) ** 2):
                path.append(6)
            if (w / (y + z) > (w / 20) ** 2) != (w / (y + 5) > (w / 100) ** 2):
                path.append(7)
            if (w / (y + z) > (w / 20) ** 2) != (w / (y + z) > (30 / 10) ** 2):
                path.append(8)
            if (w / (y + z) > (w / 20) ** 2) != (w / (y + z) > (w / 20) ** 0.6):
                path.append(9)
            if (w / (y + z) > (w / 20) ** 2) != (w / (y + z) > (w / 20) * (w / 1000)):
                path.append(10)

            if w / (y + z) > (w / 20) ** 2:
                type_code = 'C1'

            if ((y * w) % 100 + z > w / 3) != ((y * 1.8 * w) % 100 + z > w / 3):
                path.append(11)
            if ((y * w) % 100 + z > w / 3) != (((y + 20) * w) % 100 + z > w / 3):
                path.append(12)
            if ((y * w) % 100 + z > w / 3) != ((y * (w + 10)) % 100 + z > w / 3):
                path.append(13)
            if ((y * w) % 100 + z > w / 3) != ((y * w) % 90 + z > w / 3):
                path.append(14)
            if ((y * w) % 100 + z > w / 3) != ((y * w) % 100 + z * 2 > w / 3):
                path.append(15)
            if ((y * w) % 100 + z > w / 3) != ((y * w) % 100 + z > w / 4):
                path.append(16)
            if ((y * w) % 100 + z > w / 3) != ((y * w) % 100 + z > (w - 120) / 3):
                path.append(17)
            if ((y * w) % 100 + z > w / 3) != ((y * w) % 100 + 80 > w / 3):
                path.append(18)
            if ((y * w) % 100 + z > w / 3) != ((y * w) % 100 + z + 60 > w / 3):
                path.append(19)
            if ((y * w) % 100 + z > w / 3) != ((y * 100) % 100 + z > w / 3):
                path.append(20)

            if (y * w) % 100 + z > w / 3:
                type_code = 'C2'

            if (z ** 2 / w < (y - z) / 4) != ((z * 1.8) ** 2 / w < (y - z) / 4):
                path.append(21)
            if (z ** 2 / w < (y - z) / 4) != ((z + 40) ** 2 / w < (y - z) / 4):
                path.append(22)
            if (z ** 2 / w < (y - z) / 4) != (z ** 2.4 / w < (y - z) / 4):
                path.append(23)
            if (z ** 2 / w < (y - z) / 4) != (z ** 2 / w * 2.8 < (y - z) / 4):
                path.append(24)
            if (z ** 2 / w < (y - z) / 4) != (z ** 2 / w < (y - z) / 12):
                path.append(25)
            if (z ** 2 / w < (y - z) / 4) != (z ** 2 / w < (y / 2 - z) / 4):
                path.append(26)
            if (z ** 2 / w < (y - z) / 4) != (z ** 2 / w < (y - z * 2) / 4):
                path.append(27)
            if (z ** 2 / w < (y - z) / 4) != (z ** 2 / w < (y - z - 20) / 4):
                path.append(28)
            if (z ** 2 / w < (y - z) / 4) != (z ** 2 / w < (50 - z) / 4):
                path.append(29)
            if (z ** 2 / w < (y - z) / 4) != (z ** 2 / w < (y - 30) / 4):
                path.append(30)

            if z ** 2 / w < (y - z) / 4:
                type_code = 'C3'

            local_comm.send(path, dest=0, tag=21)

        elif rank == 3:
            # 接收: 气压(z), 风速(w), 日照(m)
            z = local_comm.recv(source=0, tag=1, status=status)
            w = local_comm.recv(source=0, tag=2, status=status)
            m = local_comm.recv(source=0, tag=3, status=status)

            path = []
            if (m ** 3 + w / 50 > (z + m) * 12) != ((m * 1.2) ** 3 + w / 50 > (z + m) * 12):
                path.append(1)
            if (m ** 3 + w / 50 > (z + m) * 12) != ((m + 1.8) ** 3 + w / 50 > (z + m) * 12):
                path.append(2)
            if (m ** 3 + w / 50 > (z + m) * 12) != (m ** 3.3 + w / 50 > (z + m) * 12):
                path.append(3)
            if (m ** 3 + w / 50 > (z + m) * 12) != (m ** 3 + w / 2.2 > (z + m) * 12):
                path.append(4)
            if (m ** 3 + w / 50 > (z + m) * 12) != (m ** 3 + (w + 8000) / 50 > (z + m) * 12):
                path.append(5)
            if (m ** 3 + w / 50 > (z + m) * 12) != (m ** 2 * 12 + w / 50 > (z + m) * 12):
                path.append(6)
            if (m ** 3 + w / 50 > (z + m) * 12) != (m ** 3 + 8000 / 50 > (z + m) * 12):
                path.append(7)
            if (m ** 3 + w / 50 > (z + m) * 12) != (m ** 3 + w / 50 > (z + m) * 6):
                path.append(8)
            if (m ** 3 + w / 50 > (z + m) * 12) != (m ** 3 + w / 50 > (z + 50) * 12):
                path.append(9)
            if (m ** 3 + w / 50 > (z + m) * 12) != (m ** 3 + w / 50 > (z / 2 + m) * 12):
                path.append(10)

            if m ** 3 + w / 50 > (z + m) * 12:
                type_code = 'D1'

            if ((w + z * 5) / (m + 3) > w / 8) != ((w * 1.8 + z * 5) / (m + 3) > w / 8):
                path.append(11)
            if ((w + z * 5) / (m + 3) > w / 8) != ((w + z * 10.5) / (m + 3) > w / 8):
                path.append(12)
            if ((w + z * 5) / (m + 3) > w / 8) != ((w + 150 + z * 5) / (m + 3) > w / 8):
                path.append(13)
            if ((w + z * 5) / (m + 3) > w / 8) != ((w + 18 * 5) / (m + 3) > w / 8):
                path.append(14)
            if ((w + z * 5) / (m + 3) > w / 8) != ((w + z * 5) / (m / 5 + 3) > w / 8):
                path.append(15)
            if ((w + z * 5) / (m + 3) > w / 8) != ((w + z * 5) / (m + 3) > w / 10):
                path.append(16)
            if ((w + z * 5) / (m + 3) > w / 8) != ((w + z * 5) / (m + 3) > (w - 150) / 8):
                path.append(17)
            if ((w + z * 5) / (m + 3) > w / 8) != ((w + z * 5) / (m + 3) > 15 / 8):
                path.append(18)
            if ((w + z * 5) / (m + 3) > w / 8) != ((500 + z * 5) / (m + 3) > w / 8):
                path.append(19)
            if ((w + z * 5) / (m + 3) > w / 8) != ((w + (z + 50) * 5) / (m + 3) > w / 8):
                path.append(20)

            if (w + z * 5) / (m + 3) > w / 8:
                type_code = 'D2'

            if (z * m * w / 1000 < (z + w / 100) ** 2) != (z * m * w * 10 / 1000 < (z + w / 100) ** 2):
                path.append(21)
            if (z * m * w / 1000 < (z + w / 100) ** 2) != (z * m * w / 100 + 50 < (z + w / 100) ** 2):
                path.append(22)
            if (z * m * w / 1000 < (z + w / 100) ** 2) != (z * m * (w + 5000) / 1000 < (z + w / 100) ** 2):
                path.append(23)
            if (z * m * w / 1000 < (z + w / 100) ** 2) != (z * (m + 50) * w / 1000 < (z + w / 100) ** 2):
                path.append(24)
            if (z * m * w / 1000 < (z + w / 100) ** 2) != ((z + 200) * m * w / 1000 < (z + w / 100) ** 2):
                path.append(25)
            if (z * m * w / 1000 < (z + w / 100) ** 2) != (z * m * w / 1000 < (z / 5.5 + w / 100) ** 2):
                path.append(26)
            if (z * m * w / 1000 < (z + w / 100) ** 2) != (z * m * w / 1000 < (z + w / 100) * 5):
                path.append(27)
            if (z * m * w / 1000 < (z + w / 100) ** 2) != (200 * m * w / 1000 < (z + w / 100) ** 2):
                path.append(28)
            if (z * m * w / 1000 < (z + w / 100) ** 2) != (z * 50 * w / 1000 < (z + w / 100) ** 2):
                path.append(29)
            if (z * m * w / 1000 < (z + w / 100) ** 2) != (z * m * 5000 / 1000 < (z + w / 100) ** 2):
                path.append(30)

            if z * m * w / 1000 < (z + w / 100) ** 2:
                type_code = 'D3'

            local_comm.send(path, dest=0, tag=31)


# PSO 类
class PSO:
    # --- PSO 算法参数 ---
    DIMENSION = 5  # 维度：[温度, 湿度, 气压, 风速, 日照]
    SWARM_SIZE = 20  # 粒子群大小
    MAX_ITERATIONS = 3000  # 最大迭代次数
    MIN_SPEED = -100  # 最小速度
    MAX_SPEED = 100  # 最大速度
    W = 0.729  # 惯性权重

    # 为每个维度定义位置范围
    POSITION_RANGES = [
        (1, 50),  # 图像分辨率 (百万像素)
        (10, 120),  # 采集帧率 (FPS)
        (1, 100),  # 曝光时间 (毫秒)
        (50, 1000),  # 光照强度 (lux)
        (1, 10)  # 图像清晰度 (1-10分)
    ]

    def __init__(self, local_comm, group_color):
        self.rand = random.Random()
        self.comm = local_comm
        self.group_color = group_color
        self.swarm = [self.Particle() for _ in range(PSO.SWARM_SIZE)]
        self.gBest = self.Particle()
        self.gBestFitness = float('-inf')

        self.evaluation_count = 0

    class Particle:
        def __init__(self):
            self.position = [0] * PSO.DIMENSION
            self.fitness = float('-inf')
            self.pBest = [0] * PSO.DIMENSION
            self.pBestFitness = float('-inf')
            self.velocity = [0] * PSO.DIMENSION
            self.paths = [[] for _ in range(4)]

    def initializeSwarm(self):
        """初始化粒子群，使其在指定范围内生成"""
        samplesPerDimension = self.SWARM_SIZE
        for dim in range(self.DIMENSION):
            rangeMin, rangeMax = self.POSITION_RANGES[dim]

            samples = []
            intervalSize = (rangeMax - rangeMin) / samplesPerDimension
            for i in range(samplesPerDimension):
                start = rangeMin + i * intervalSize
                sample = start + self.rand.random() * intervalSize
                samples.append(sample)

            self.rand.shuffle(samples)
            for i in range(self.SWARM_SIZE):
                self.swarm[i].position[dim] = samples[i]
                self.swarm[i].velocity[dim] = self.MIN_SPEED + self.rand.random() * (self.MAX_SPEED - self.MIN_SPEED)

    def updateParticles(self, currentTargetPaths):
        """更新粒子群中所有粒子的速度、位置、适应度以及pBest和gBest"""
        for particle in self.swarm:
            # 1. 更新速度和位置
            for i in range(PSO.DIMENSION):
                r1 = self.rand.random()
                r2 = self.rand.random()
                particle.velocity[i] = self.W * particle.velocity[i] + \
                                       2.0 * r1 * (particle.pBest[i] - particle.position[i]) + \
                                       2.0 * r2 * (self.gBest.position[i] - particle.position[i])
                particle.velocity[i] = max(PSO.MIN_SPEED, min(PSO.MAX_SPEED, particle.velocity[i]))
                particle.position[i] += particle.velocity[i]

                rangeMin, rangeMax = self.POSITION_RANGES[i]
                particle.position[i] = max(rangeMin, min(rangeMax, particle.position[i]))

            # 2. 计算新位置的适应度
            mpi_result = self.generateMPIPaths(particle.position)
            particle.paths = mpi_result
            fitness = self.calculateFitnessBasedOnPaths(mpi_result, currentTargetPaths)
            particle.fitness = fitness

            # 3. 更新个体最优 (pBest)
            if fitness > particle.pBestFitness:
                particle.pBest = particle.position.copy()
                particle.pBestFitness = fitness

            # 4. 更新全局最优 (gBest)
            if fitness > self.gBestFitness:
                self.gBest = copy.deepcopy(particle)
                self.gBestFitness = fitness

    def generateMPIPaths(self, a):
        """主进程与工作进程通信，处理5个变量"""
        comm = self.comm
        status = MPI.Status()
        path = []
        x, y, z, w, m = a[0], a[1], a[2], a[3], a[4]

        comm.send(x, dest=1, tag=1)
        comm.send(y, dest=1, tag=2)
        comm.send(z, dest=1, tag=3)

        comm.send(y, dest=2, tag=1)
        comm.send(z, dest=2, tag=2)
        comm.send(w, dest=2, tag=3)

        comm.send(z, dest=3, tag=1)
        comm.send(w, dest=3, tag=2)
        comm.send(m, dest=3, tag=3)

        # 主进程的路径计算
        if ((x + y + z) ** 2 > w * m * 5) != ((x * 2 + y + z) ** 2 > w * m * 5):
            path.append(1)
        if ((x + y + z) ** 2 > w * m * 5) != ((x + y * 1.8 + z) ** 2 > w * m * 5):
            path.append(2)
        if ((x + y + z) ** 2 > w * m * 5) != ((x + y + z * 1.8) ** 2 > w * m * 5):
            path.append(3)
        if ((x + y + z) ** 2 > w * m * 5) != ((x + y + z) ** 2.1 > w * m * 5):
            path.append(4)
        if ((x + y + z) ** 2 > w * m * 5) != ((x + y + z) ** 2 > w * m * 3):
            path.append(5)
        if ((x + y + z) ** 2 > w * m * 5) != ((x + y + z) ** 2 > w * (m - 2) * 5):
            path.append(6)
        if ((x + y + z) ** 2 > w * m * 5) != ((x + y + z) ** 2 > (w - 180) * m * 5):
            path.append(7)
        if ((x + y + z) ** 2 > w * m * 5) != ((x + y + z) ** 2 > w * 4.5 * 5):
            path.append(8)
        if ((x + y + z) ** 2 > w * m * 5) != ((x + y + 20) ** 2 > w * m * 5):
            path.append(9)
        if ((x + y + z) ** 2 > w * m * 5) != ((x + 100 + z) ** 2 > w * m * 5):
            path.append(10)

        if (x + y + z) ** 2 > w * m * 5:
            type_code = 'A1'

        if (x ** 4 / 100 > y * z) != ((x + 8) ** 4 / 100 > y * z):
            path.append(11)
        if (x ** 4 / 100 > y * z) != ((x * 1.5) ** 4 / 100 > y * z):
            path.append(12)
        if (x ** 4 / 100 > y * z) != (x ** 4.8 / 100 > y * z):
            path.append(13)
        if (x ** 4 / 100 > y * z) != (x ** 4 / 30 > y * z):
            path.append(14)
        if (x ** 4 / 100 > y * z) != (x ** 4 / 100 > y * z / 2.5):
            path.append(15)
        if (x ** 4 / 100 > y * z) != (x ** 4 / 100 > y * (z - 30)):
            path.append(16)
        if (x ** 4 / 100 > y * z) != (x ** 4 / 100 > (y - 30) * z):
            path.append(17)
        if (x ** 4 / 100 > y * z) != (x ** 4 / 100 + 1800 > y * z):
            path.append(18)
        if (x ** 4 / 100 > y * z) != (x ** 4 / 100 > 20 * z):
            path.append(19)
        if (x ** 4 / 100 > y * z) != (x ** 4 / 100 > y * 15):
            path.append(20)

        if x ** 4 / 100 > y * z:
            type_code = 'A2'

        if ((x * y) / (z + w / 100) > m * 15) != ((x * 1.6 * y) / (z + w / 100) > m * 15):
            path.append(21)
        if ((x * y) / (z + w / 100) > m * 15) != ((x * (y + 50)) / (z + w / 100) > m * 15):
            path.append(22)
        if ((x * y) / (z + w / 100) > m * 15) != (((x + 20) * y) / (z + w / 100) > m * 15):
            path.append(23)
        if ((x * y) / (z + w / 100) > m * 15) != ((x * y) / (z / 1.8 + w / 100) > m * 15):
            path.append(24)
        if ((x * y) / (z + w / 100) > m * 15) != ((x * y) / (z + w / 100) + 20 > m * 15):
            path.append(25)
        if ((x * y) / (z + w / 100) > m * 15) != ((x * y) / (z + w / 100) > m * 9):
            path.append(26)
        if ((x * y) / (z + w / 100) > m * 15) != ((x * y) / (z + w / 100) > (m - 2) * 15):
            path.append(27)
        if ((x * y) / (z + w / 100) > m * 15) != ((x * y) / (z + w / 100) > 18 * 15):
            path.append(28)
        if ((x * y) / (z + w / 100) > m * 15) != ((x * 50) / (z + w / 100) > m * 15):
            path.append(29)
        if ((x * y) / (z + w / 100) > m * 15) != ((50 * y) / (z + w / 100) > m * 15):
            path.append(30)

        if (x * y) / (z + w / 100) > m * 15:
            type_code = 'A3'

        path1 = comm.recv(source=1, tag=11, status=status)
        path2 = comm.recv(source=2, tag=21, status=status)
        path3 = comm.recv(source=3, tag=31, status=status)

        return [path, path1, path2, path3]

    def calculateFitnessBasedOnPaths(self, mpi_paths, targetPaths):
        """计算生成的路径与目标路径的匹配度作为适应度"""
        self.evaluation_count += 1

        total_fitness = 0.0
        for i in range(4):
            target = targetPaths[i]
            if target:
                match_count = sum(1 for branch in target if branch in mpi_paths[i])
                local_fitness = match_count / len(target)
            else:
                local_fitness = 1.0 if not mpi_paths[i] else 0.0
            total_fitness += local_fitness
        return total_fitness / 4.0

    def run(self, currentTargetPaths):
        """执行PSO算法的主循环，返回结果"""
        start_time = MPI.Wtime()
        found = False
        actual_iterations = 0

        for iteration in range(PSO.MAX_ITERATIONS):
            self.updateParticles(currentTargetPaths)
            actual_iterations = iteration + 1

            if self.gBestFitness == 1.0:
                found = True
                break

        end_time = MPI.Wtime()
        elapsed = end_time - start_time
        return found, elapsed, actual_iterations, self.evaluation_count


def save_all_results_to_excel(all_group_results):
    """将所有组的实验结果保存到Excel文件中"""
    output_dir = "experimental_results"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    num_experiments = len(all_group_results[0]['success'])
    num_groups = len(all_group_results)

    file_types = [
        ("success", "是否成功"),
        ("iterations", "迭代次数"),
        ("evaluations", "评价次数"),
        ("time", "迭代时间"),
        ("fitness", "最终适应度")
    ]

    for file_type, description in file_types:
        wb = Workbook()
        ws = wb.active
        ws.title = description

        ws.cell(row=1, column=1, value="实验次数")
        for group_id in range(num_groups):
            ws.cell(row=1, column=group_id + 2, value=f"路径{group_id + 1}")

        for exp_id in range(num_experiments):
            ws.cell(row=exp_id + 2, column=1, value=exp_id + 1)
            for group_id in range(num_groups):
                value = all_group_results[group_id][file_type][exp_id]
                ws.cell(row=exp_id + 2, column=group_id + 2, value=value)

        filename = f"{output_dir}/{file_type}.xlsx"
        wb.save(filename)
        print(f"[全局] 已保存 {description} 数据到: {filename}")


def print_statistics(all_group_results, num_groups):
    """打印统计结果"""
    print("\n" + "=" * 80)
    print("实验统计结果".center(80))
    print("=" * 80)

    print("\n" + "-" * 80)
    print("各组统计".center(80))
    print("-" * 80)
    print(f"{'组ID':<8} {'成功率':<12} {'平均迭代':<12} {'平均评价':<12} {'平均时间':<12} {'平均适应度':<12}")
    print("-" * 80)

    group_stats = []
    for group_id in range(num_groups):
        results = all_group_results[group_id]
        success_count = sum(results['success'])
        total_count = len(results['success'])

        avg_iterations = sum(results['iterations']) / total_count
        avg_evaluations = sum(results['evaluations']) / total_count
        avg_time = sum(results['time']) / total_count
        avg_fitness = sum(results['fitness']) / total_count

        success_rate = success_count / total_count

        print(
            f"组{group_id:<6} {success_count}/{total_count:<9} {avg_iterations:<12.1f} {avg_evaluations:<12.0f} {avg_time:<12.4f}s {avg_fitness:<12.6f}")

        group_stats.append({
            'group_id': group_id,
            'success_rate': success_rate,
            'success_count': success_count,
            'total_count': total_count,
            'avg_iterations': avg_iterations,
            'avg_evaluations': avg_evaluations,
            'avg_time': avg_time,
            'avg_fitness': avg_fitness
        })

    print("\n" + "=" * 80)
    print("全局汇总".center(80))
    print("=" * 80)

    all_success = []
    all_iterations = []
    all_evaluations = []
    all_times = []
    all_fitness = []

    for group_results in all_group_results:
        all_success.extend(group_results['success'])
        all_iterations.extend(group_results['iterations'])
        all_evaluations.extend(group_results['evaluations'])
        all_times.extend(group_results['time'])
        all_fitness.extend(group_results['fitness'])

    total_experiments = len(all_success)
    total_success = sum(all_success)
    total_fail = total_experiments - total_success

    print(f"\n总实验次数: {total_experiments} ({num_groups}组 × {NUM_EXPERIMENTS}次)")
    print(f"总成功次数: {total_success} ({total_success / total_experiments * 100:.2f}%)")
    print(f"总失败次数: {total_fail} ({total_fail / total_experiments * 100:.2f}%)")

    print(f"\n迭代次数统计:")
    print(f"  平均: {sum(all_iterations) / len(all_iterations):.1f}")
    print(f"  中位数: {sorted(all_iterations)[len(all_iterations) // 2]:.1f}")
    print(f"  最小: {min(all_iterations)}")
    print(f"  最大: {max(all_iterations)}")

    print(f"\n评价次数统计:")
    print(f"  平均: {sum(all_evaluations) / len(all_evaluations):.0f}")
    print(f"  中位数: {sorted(all_evaluations)[len(all_evaluations) // 2]:.0f}")
    print(f"  最小: {min(all_evaluations)}")
    print(f"  最大: {max(all_evaluations)}")

    print(f"\n运行时间统计:")
    print(f"  平均: {sum(all_times) / len(all_times):.4f}s")
    print(f"  中位数: {sorted(all_times)[len(all_times) // 2]:.4f}s")
    print(f"  最小: {min(all_times):.4f}s")
    print(f"  最大: {max(all_times):.4f}s")

    print(f"\n适应度统计:")
    print(f"  平均: {sum(all_fitness) / len(all_fitness):.6f}")
    print(f"  中位数: {sorted(all_fitness)[len(all_fitness) // 2]:.6f}")
    print(f"  最小: {min(all_fitness):.6f}")
    print(f"  最大: {max(all_fitness):.6f}")

    sorted_by_success = sorted(group_stats, key=lambda x: x['success_rate'])
    print(
        f"\n最难优化的组: 组{sorted_by_success[0]['group_id']} (成功率 {sorted_by_success[0]['success_rate'] * 100:.1f}%)")
    print(
        f"最易优化的组: 组{sorted_by_success[-1]['group_id']} (成功率 {sorted_by_success[-1]['success_rate'] * 100:.1f}%)")

    print("\n" + "=" * 80)


def main():
    global_start = MPI.Wtime()

    targetPathsSets = [
        [[1, 2, 3, 4, 5, 7, 8, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20],
         [1, 2, 3, 4, 5, 6, 7, 9, 10, 11, 13, 14, 15, 17, 19, 20, 21, 22, 24, 25, 27, 29, 30],
         [12, 15, 16, 17, 18, 19, 22, 26], [11, 12, 13, 15, 16, 17, 18, 19, 20, 22, 23, 26, 30]

         ],

        [[1, 2, 4, 5, 6, 7, 8, 9, 10, 21, 22, 23, 24, 25, 26, 27, 29, 30], [5, 6, 7, 9, 22, 25, 27, 29, 30],
         [1, 2, 3, 4, 7, 9, 10, 11, 13, 17, 18, 19, 22],
         [9, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 25, 28, 30]

         ],

        [[1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 21, 22, 23, 24, 25, 26, 27, 29], [1, 4, 5, 6, 7, 9, 10, 25],
         [1, 2, 4, 7, 9, 10, 17, 18, 19, 22, 26, 27, 28, 30],
         [9, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 25, 28, 30]

         ],

        [[1, 2, 4, 5, 6, 7, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 21, 22, 23, 24, 25, 26, 27, 30],
         [6, 9, 11, 13, 14, 19, 20, 22, 25], [1, 4, 10, 13, 14, 20, 22],
         [1, 2, 3, 4, 5, 6, 7, 8, 10, 22, 23, 25, 28, 30]

         ],

        [[1, 2, 4, 5, 6, 7, 8, 9, 10, 21, 22, 23, 24, 25, 26, 27, 30],
         [1, 2, 3, 4, 5, 6, 7, 9, 10, 12, 16, 18, 22, 25, 27, 29, 30], [10, 17, 18, 19],
         [11, 13, 15, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 28, 30]

         ],

        [[2, 8, 10, 11, 12, 13, 14, 15, 16, 18, 19], [1, 2, 3, 4, 5, 6, 7, 9, 10, 11, 13, 14, 21, 22, 25, 27],
         [15, 16, 17, 18, 19, 21, 22, 23, 24, 25, 26, 27, 28], [9, 11, 12, 13, 15, 16, 17, 18, 20, 23, 26, 30]

         ],

        [[2, 5, 6, 7, 9, 10, 11, 12, 13, 14, 16, 17, 18, 21, 22, 23, 26, 27, 29, 30],
         [6, 13, 14, 19, 20, 21, 22, 25, 27, 28, 29, 30], [1, 2, 3, 4, 5, 7, 9, 10, 11, 13, 20, 22],
         [9, 22, 23, 24, 25, 28, 30]

         ],

        [[2, 3, 4, 5, 8, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20], [11, 13, 14, 15, 17, 18, 20],
         [21, 22, 23, 24, 25, 26, 27, 28, 29], [1, 2, 3, 4, 5, 6, 7, 8, 10, 11, 12, 13, 15, 16, 17, 18, 20, 26]

         ],

        [[7, 10, 20, 22, 29, 30], [11, 13, 14, 15, 19, 20, 22, 25, 27, 29, 30],
         [1, 2, 3, 4, 5, 6, 7, 9, 10, 11, 12, 20, 22, 28, 30],
         [11, 13, 14, 15, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 28, 30]

         ],

        [[11, 12, 13, 14, 15, 16, 17, 18, 19, 20], [1, 2, 3, 4, 6, 7, 9, 10, 11, 13, 14, 15, 20, 23, 24, 26, 28, 29],
         [8, 9, 21, 22, 23, 24, 25, 26, 27, 28], [9, 11, 15, 18, 21, 22, 23, 25, 26, 27]

         ],
    ]

    comm_world = MPI.COMM_WORLD
    size = comm_world.Get_size()
    rank_world = comm_world.Get_rank()

    # 串行执行模式：只需要 4 个进程
    if size != 4:
        if rank_world == 0:
            print(f"错误：串行模式需要且仅需 4 个 MPI 进程（1主3从），当前进程数为 {size}。")
        MPI.Finalize()
        exit()

    local_comm = comm_world
    local_rank = rank_world
    num_paths = len(targetPathsSets)

    if rank_world == 0:
        print("\n" + "=" * 80)
        print("串行实验配置".center(80))
        print("=" * 80)
        print(f"进程总数: {size} (1个主进程 + 3个工作进程)")
        print(f"总路径集数: {num_paths}")
        print(f"每组实验次数: {NUM_EXPERIMENTS}")
        print(f"PSO参数: 粒子数={PSO.SWARM_SIZE}, 最大迭代={PSO.MAX_ITERATIONS}")
        print("=" * 80)
        print(f"\n[运行中] 正在依次执行 {num_paths} 组路径实验...\n")

        all_results = []

        # 外层循环：逐个路径集进行实验
        for set_idx in range(num_paths):
            currentTargetPaths = targetPathsSets[set_idx]
            print(f"[进度] 正在执行路径集 {set_idx + 1}/{num_paths}...")

            success_list = []
            iteration_list = []
            evaluations_list = []
            time_list = []
            fitness_list = []

            for exp_id in range(NUM_EXPERIMENTS):
                # 保持随机种子与原逻辑一致，但通过 set_idx 区分
                random.seed(set_idx * 10000 + exp_id)

                psoInstance = PSO(local_comm, group_color=set_idx)
                psoInstance.initializeSwarm()
                found, elapsed, actual_iterations, evaluations = psoInstance.run(currentTargetPaths)

                success_list.append(1 if found else 0)
                iteration_list.append(actual_iterations)
                evaluations_list.append(evaluations)
                time_list.append(elapsed)
                fitness_list.append(psoInstance.gBestFitness)

            # 存储该路径集的所有结果
            all_results.append({
                'success': success_list,
                'iterations': iteration_list,
                'evaluations': evaluations_list,
                'time': time_list,
                'fitness': fitness_list
            })

        # 所有路径循环结束后，通知 worker 进程退出
        for dest in range(1, 4):
            local_comm.send("QUIT", dest=dest, tag=0)

        # 打印统计并保存
        print_statistics(all_results, num_paths)
        save_all_results_to_excel(all_results)

        global_end = MPI.Wtime()
        print(f"\n程序总运行时间：{global_end - global_start:.4f} 秒\n")

    else:
        # 工作进程逻辑：持续待命，直到主进程发来 QUIT
        worker_loop(local_comm)

    MPI.Finalize()


if __name__ == "__main__":
    main()