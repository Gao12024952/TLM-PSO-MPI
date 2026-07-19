import random
import time
import copy
import math  # 添加math模块用于计算标准差
from mpi4py import MPI
from collections import Counter
import openpyxl  # 新增：用于Excel文件操作
from openpyxl import Workbook
import os  # 新增：用于文件路径操作

# 实验配置
NUM_EXPERIMENTS = 25  # 实验次数
WARMUP_EXPERIMENTS = 5  # 前10次实验作为预热

# 分成局部和全局两个迭代次数控制
GENL = 4  # 局部粒子迭代周期
GENO = 5  # 全局粒子迭代周期
MAX_ITERATIONS = 3000
OPTIMAL_FITNESS = 1.0
H = 2  # 全局粒子反馈给局部的数量

# 定义5维参数的物理范围
POSITION_RANGES = [
    (100, 5000),  # 用电负荷 (100-5000 MW)
    (30, 95),  # 节能效率 (30-95%)
    (1, 12),  # 峰值持续时间 (1-12 小时)
    (5, 70),  # 可再生能源比例 (5-70%)
    (1, 10)  # 设备运行状态 (1-10分)
]

# 修正：恢复为4条路径
targetPathsSets = [
    [[1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 21, 22, 23, 24, 25, 27, 28, 29, 30],
     [9, 11, 13, 15, 17, 18, 21, 23, 28, 29],
     [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 17, 18, 19, 21, 24, 26, 27, 29],
     [1, 2, 3, 4, 6, 7, 8, 9, 10, 12, 14, 19, 22, 24, 25, 27, 30]

     ],

    [[1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 26],
     [9, 11, 12, 13, 15, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30],
     [17, 18, 19, 21, 24, 26, 27, 29],
     [1, 2, 3, 4, 6, 7, 8, 9, 10, 12, 14, 16, 19, 21, 22, 23, 25, 27, 28, 29, 30]

     ],

    [[1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 22, 24, 25, 27, 28, 29],
     [9, 12, 16, 19, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30],
     [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 17, 18, 19, 21, 24, 26, 27, 29],
     [4, 12, 14, 19, 21, 25, 27]

     ],

    [[1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 15, 16, 17, 18, 22, 24, 25, 27, 28, 29, 30],
     [9, 17, 21, 22, 23, 24, 27, 28, 29, 30],
     [2, 3, 4, 5, 6, 7, 8, 9, 17, 18, 19, 21, 23, 24, 26, 27, 28, 29, 30],
     [4, 12, 14, 19, 22, 24, 26, 27, 28]

     ],

    [[2, 4, 5, 6, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 24, 25, 27, 29],
     [9, 12, 18, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30],
     [17, 18, 19, 26],
     [1, 2, 3, 4, 6, 8, 9, 10, 11, 12, 13, 14, 15, 16, 19, 20, 21, 22, 23, 24, 25, 28, 29, 30]

     ],

    [[11, 13, 14, 15, 16, 17, 18, 22, 24, 26, 27, 29],
     [4, 6, 7, 10, 11, 12, 21, 22, 23, 24, 25, 26, 27, 28, 30],
     [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 23, 28, 29, 30],
     [1, 2, 3, 4, 6, 7, 8, 9, 10, 17, 19, 21, 23, 24, 25, 27, 28, 29, 30]

     ],
    [[2, 4, 5, 6, 8, 9, 10, 11, 13, 15, 16, 17, 18, 22, 24, 26, 27, 29, 30],
     [4, 5, 6, 7, 10, 12, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30],
     [11, 14, 23, 26, 28],
     [5, 11, 13, 15, 16, 17, 19, 20, 21, 22, 23, 24, 25, 27, 28, 29, 30]

     ],

    [[24, 25, 28, 29],
     [9, 12, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30],
     [1, 2, 3, 4, 6, 7, 8, 9, 10, 17, 18, 19, 21, 22, 24, 27, 29],
     [1, 2, 3, 4, 6, 7, 8, 9, 10, 12, 14, 19, 21, 22, 23, 24, 25, 28, 29, 30]

     ],

    [[1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 19, 20, 24, 25],
     [9, 11, 12, 13, 15, 18, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30],
     [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 17, 18, 19, 21, 24, 26, 27, 29],
     [4, 12, 14, 19, 22, 24, 26, 27, 28]

     ],

    [[1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 21, 22, 23, 24, 25, 27, 28, 29, 30],
     [9, 11, 13, 14, 15, 16, 17, 19],
     [17, 18, 19, 26, 27, 29],
     [1, 2, 3, 4, 6, 7, 8, 9, 10, 12, 14, 16, 19, 21, 22, 23, 25, 27, 28, 29, 30]

     ],

    [[1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 15, 16, 17, 18, 25],
     [9, 20, 21, 23, 24, 27, 28, 29],
     [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 17, 18, 19, 21, 24, 26, 27, 29],
     [12, 14, 18, 22, 24, 25, 27, 29, 30]

     ],

    [[24],
     [3, 4, 5, 6, 7, 10, 12, 16, 19, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30],
     [1, 2, 3, 4, 6, 7, 8, 9, 10, 12, 13, 20, 22, 27, 29],
     [1, 2, 3, 4, 6, 7, 8, 9, 10, 19, 21, 22, 23, 25, 27, 28, 29, 30]

     ],

    [[13, 15, 16, 17, 18, 21, 22, 23, 24, 25, 27, 28, 29, 30],
     [4, 5, 6, 7, 8, 10, 11],
     [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 21, 23, 24, 26, 28, 29, 30],
     [1, 3, 4, 6, 19, 21, 22, 23, 24, 25, 27, 28, 29, 30]

     ],

    [[9, 11, 12, 13, 14, 15, 16, 17, 18, 21, 22, 23, 24, 25, 27, 28, 29, 30],
     [1, 2, 3, 4, 5, 6, 7, 8, 10, 13, 14, 19],
     [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 29],
     [3, 4, 21, 22, 23, 24, 25, 27, 28, 29, 30]

     ],

    [[11, 12, 13, 15, 16, 17, 18],
     [11, 12, 13, 15, 18, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30],
     [4, 6, 7, 8, 9, 12, 13, 15, 20, 22, 27, 29, 30],
     [7, 22, 26, 29]

     ],

    [[7, 11, 12, 13, 15, 16, 17, 18],
     [12, 13, 15, 16, 19, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30],
     [11, 12, 13, 14, 15, 16, 20],
     [9, 21, 22, 23, 29, 30]

     ],
]


# 添加标准差计算函数
def calculate_std_dev(values):
    """计算标准差"""
    if len(values) <= 1:
        return 0.0
    mean = sum(values) / len(values)
    variance = sum((x - mean) ** 2 for x in values) / (len(values) - 1)  # 样本标准差
    return math.sqrt(variance)


# 🔧 修改：新增Excel输出功能 - 将所有路径组的结果合并保存到Excel文件，包含迭代次数
def save_all_results_to_excel(all_group_results, num_path_sets):
    """将所有组的实验结果保存到Excel文件中，每种类型一个文件，所有路径作为列"""

    # 创建输出目录（如果不存在）
    output_dir = "experimental_results"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    if not all_group_results or len(all_group_results) == 0:
        print("[Excel输出] 警告：没有数据可保存")
        return

    # 获取实验次数（从第一个有数据的组获取）
    num_experiments = 0
    for group_data in all_group_results.values():
        if 'success_list' in group_data and len(group_data['success_list']) > 0:
            num_experiments = len(group_data['success_list'])
            break

    if num_experiments == 0:
        print("[Excel输出] 警告：所有组都没有实验数据")
        return

    # 🔧 修改：定义4种类型的数据，包含迭代次数
    file_types = [
        ("success", "是否成功"),
        ("time", "迭代时间"),
        ("fitness", "最终适应度"),
        ("iterations", "迭代次数"),  # ← 新增
        ("evaluations", "评价次数")  # 🔧 新增:评价次数
    ]

    for file_type, description in file_types:
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = description

        # 设置表头 - 第一列是实验次数，其余列是各路径
        ws.cell(row=1, column=1, value="实验次数")
        for group_id in range(num_path_sets):
            ws.cell(row=1, column=group_id + 2, value=f"路径{group_id + 1}")

        # 填入数据
        for exp_id in range(num_experiments):
            ws.cell(row=exp_id + 2, column=1, value=exp_id + 1)  # 实验次数
            for group_id in range(num_path_sets):
                if group_id in all_group_results:
                    group_data = all_group_results[group_id]
                    data_key = f"{file_type}_list"
                    if data_key in group_data and exp_id < len(group_data[data_key]):
                        value = group_data[data_key][exp_id]
                        ws.cell(row=exp_id + 2, column=group_id + 2, value=value)
                    else:
                        # 🔧 修改：如果没有数据，填入默认值，包含迭代次数的处理
                        if file_type == "success":
                            ws.cell(row=exp_id + 2, column=group_id + 2, value=0)
                        elif file_type == "time":
                            ws.cell(row=exp_id + 2, column=group_id + 2, value=0)
                        elif file_type == "iterations":  # ← 新增
                            ws.cell(row=exp_id + 2, column=group_id + 2, value=0)
                        elif file_type == "evaluations":  # 🔧 新增:评价次数默认值处理
                            ws.cell(row=exp_id + 2, column=group_id + 2, value=0)
                        else:  # fitness
                            ws.cell(row=exp_id + 2, column=group_id + 2, value=0.0)
                else:
                    # 🔧 修改：如果组不存在，填入默认值，包含迭代次数的处理
                    if file_type == "success":
                        ws.cell(row=exp_id + 2, column=group_id + 2, value=0)
                    elif file_type == "time":
                        ws.cell(row=exp_id + 2, column=group_id + 2, value=0)
                    elif file_type == "iterations":  # ← 新增
                        ws.cell(row=exp_id + 2, column=group_id + 2, value=0)
                    elif file_type == "evaluations":  # 🔧 新增:评价次数默认值处理
                        ws.cell(row=exp_id + 2, column=group_id + 2, value=0)
                    else:  # fitness
                        ws.cell(row=exp_id + 2, column=group_id + 2, value=0.0)

        # 保存文件
        filename = f"{output_dir}/{file_type}.xlsx"
        wb.save(filename)
        print(f"[Excel输出] 已保存 {description} 数据到: {filename}")


def set_experiment_seeds(experiment_id, rank):
    """为每次实验设置独立的随机种子"""
    seed = hash((experiment_id, rank)) % (2 ** 32)
    random.seed(seed)


def reset_communication_state(comm):
    """重置MPI通信状态"""
    comm.Barrier()  # 确保所有进程同步


# 相似度计算函数
def calculate_similarity(path1, path2):
    if not path1 or not path2:
        return 0

    # 使用Counter计数器统计路径中元素出现次数
    path1_counter = Counter(path1)
    path2_counter = Counter(path2)

    # 计算匹配元素数量
    matched = 0
    for code, count in path1_counter.items():
        matched += min(count, path2_counter.get(code, 0))

    # 计算相似度
    return matched / max(len(path1), len(path2))


# --- 局部三维PSO类（用于进程1、2、3的局部更新） ---
class PSO:
    DIMENSION = 3
    SWARM_SIZE = 5
    MAX_ITERATIONS = 8000
    MIN_SPEED = -100
    MAX_SPEED = 100
    W = 0.729
    C1 = 2.0
    C2 = 2.0
    COMM_FREQUENCY = 3  # 局部配对粒子通信频率

    class Particle:
        def __init__(self, dimension):
            self.position = [0] * dimension  # 位置
            self.velocity = [0] * dimension  # 速度
            self.fitness = float('-inf')  # 当前适应度
            self.pBest = [0] * dimension  # 个人历史最优位置
            self.pBestFitness = float('-inf')  # 个人历史最优适应度
            self.last_calculated_fitness = None  # 最近一次计算的适应度
            self.paths = None  # 用于存储路径数据

    def __init__(self, mode, target_path, global_comm=None, partner_rank=None,
                 group_color=0, local_rank=0):
        self.mode = mode  # "full" 或 "partial" 或 "path3"
        self.target_path = target_path  # 局部目标路径（适应度计算依据）
        self.swarm = [PSO.Particle(PSO.DIMENSION) for _ in range(PSO.SWARM_SIZE)]
        self.gBest = PSO.Particle(PSO.DIMENSION)  # 全局最优粒子（局部解中的最优）
        self.gBestFitness = float('-inf')

        # 添加当前代最优粒子属性
        self.currentGenBest = PSO.Particle(PSO.DIMENSION)  # 当前代最优粒子
        self.currentGenBestFitness = float('-inf')  # 当前代最优适应度

        self.rand = random.Random()
        self.iteration = 0  # 当前迭代次数
        self.evaluation_count = 0  # 🔧 新增:评价次数计数器

        # 根据模式确定维度映射
        if mode == "full":  # xyz -> 温度、湿度、气压 -> 维度0,1,2
            self.dimension_indices = [0, 1, 2]
        elif mode == "partial":  # yzw -> 湿度、气压、风速 -> 维度1,2,3
            self.dimension_indices = [1, 2, 3]
        elif mode == "path3":  # zwm -> 气压、风速、日照 -> 维度2,3,4
            self.dimension_indices = [2, 3, 4]
        else:
            self.dimension_indices = [0, 1, 2]  # 默认值

        # 跨组通信相关属性（随机配对）
        self.global_comm = global_comm  # 全局MPI通信器
        self.partner_rank = partner_rank  # 随机配对伙伴进程的rank
        self.partner_gBest = None  # 伙伴的全局最优位置
        self.partner_fitness = float('-inf')  # 伙伴的最优适应度

        self.comm_time = 0.0  # 通信时间计数器
        self.partner_finished = False  # 标志随机配对伙伴是否已完成
        self.global_solution_found = False  # 标志全局优化是否已找到最优解
        self.group_color = group_color  # 组颜色/ID，用于日志
        self.local_rank = local_rank  # 当前进程在组内的编号

        self.initializeSwarm()

    def initializeSwarm(self):
        samplesPerDimension = PSO.SWARM_SIZE
        for d in range(PSO.DIMENSION):
            # 获取当前维度对应的全局维度索引
            global_dim = self.dimension_indices[d]
            rangeMin, rangeMax = POSITION_RANGES[global_dim]

            samples = []
            intervalSize = (rangeMax - rangeMin) / samplesPerDimension
            for i in range(samplesPerDimension):
                start = rangeMin + i * intervalSize
                sample = start + self.rand.random() * intervalSize
                samples.append(sample)
            self.rand.shuffle(samples)

            for i in range(PSO.SWARM_SIZE):
                self.swarm[i].position[d] = int(round(samples[i]))
                # 速度初始化使用相对范围
                velocity_range = (rangeMax - rangeMin) / 10  # 使用范围的1/10作为速度范围
                self.swarm[i].velocity[d] = int(round(-velocity_range + self.rand.random() * 2 * velocity_range))
                self.swarm[i].pBest[d] = self.swarm[i].position[d]

        for particle in self.swarm:
            particle.fitness = self.calculateFitnessBasedOnPath(particle.position)
            particle.last_calculated_fitness = particle.fitness
            particle.pBestFitness = particle.fitness
            if particle.fitness >= self.gBestFitness:
                self.gBestFitness = particle.fitness
                self.gBest = copy.deepcopy(particle)

    def notifyPartnerOfGlobalTermination(self):
        """当全局优化找到最优解时，通知随机配对伙伴进程"""
        if self.partner_rank is not None and self.global_comm is not None and not self.partner_finished:
            start_comm = time.time()
            req = self.global_comm.isend("finished", dest=self.partner_rank, tag=100)
            req.Wait()
            end_comm = time.time()
            self.comm_time += (end_comm - start_comm)
            self.partner_finished = True  # 设置为已通知

    def updateParticles(self):
        self.iteration += 1  # 增加迭代计数

        # 重置当前代最优粒子
        self.currentGenBestFitness = float('-inf')
        self.currentGenBest = PSO.Particle(PSO.DIMENSION)

        # 检查是否全局找到最优解，如果是，则通知伙伴并停止更新
        if self.global_solution_found:
            if not self.partner_finished:
                self.notifyPartnerOfGlobalTermination()
            return

        # 1. 【配对接收】检查来自随机配对伙伴的消息（非阻塞）
        if self.partner_rank is not None and self.global_comm is not None and not self.partner_finished:
            while self.global_comm.Iprobe(source=self.partner_rank, tag=100):
                start_comm = time.time()
                partner_data = self.global_comm.recv(source=self.partner_rank, tag=100)
                end_comm = time.time()
                self.comm_time += (end_comm - start_comm)

                # 检查伙伴是否已完成
                if partner_data == "finished":
                    self.partner_finished = True
                    self.partner_gBest = None
                    self.partner_fitness = float('-inf')
                else:
                    # 计算接收到的位置的适应度
                    new_partner_fitness = self.calculateFitnessBasedOnPath(partner_data)
                    if new_partner_fitness >= self.partner_fitness and new_partner_fitness != 0:
                        self.partner_gBest = partner_data
                        self.partner_fitness = new_partner_fitness

        # 2. 确定是否使用配对伙伴信息
        use_partner = False
        if self.partner_gBest is not None:
            partner_fitness = self.partner_fitness
            # 直接进行适应度判断，去掉概率性接受
            if self.gBestFitness == float('-inf'):
                use_partner = True
            else:
                if partner_fitness > self.gBestFitness or abs(self.gBestFitness - partner_fitness) < 0.3:
                    use_partner = True

        # 3. 【核心计算】更新粒子
        for particle in self.swarm:
            for i in range(PSO.DIMENSION):
                r1 = self.rand.random()
                r2 = self.rand.random()

                # 修改后的速度更新，包括伙伴信息（如果有）
                if use_partner:
                    r3 = self.rand.random()
                    particle.velocity[i] = (PSO.W * particle.velocity[i] +
                                            PSO.C1 * r1 * (particle.pBest[i] - particle.position[i]) +
                                            PSO.C2 * r2 * (self.gBest.position[i] - particle.position[i]) +
                                            PSO.C1 * r3 * (self.partner_gBest[i] - particle.position[i]))
                else:
                    particle.velocity[i] = (PSO.W * particle.velocity[i] +
                                            PSO.C1 * r1 * (particle.pBest[i] - particle.position[i]) +
                                            PSO.C2 * r2 * (self.gBest.position[i] - particle.position[i]))

                particle.velocity[i] = max(PSO.MIN_SPEED, min(PSO.MAX_SPEED, particle.velocity[i]))
                particle.position[i] += int(round(particle.velocity[i]))

                # 使用对应维度的边界约束
                global_dim = self.dimension_indices[i]
                rangeMin, rangeMax = POSITION_RANGES[global_dim]
                particle.position[i] = int(round(max(rangeMin, min(rangeMax, particle.position[i]))))

            # 计算适应度并更新个人最优
            particle.fitness = self.calculateFitnessBasedOnPath(particle.position)
            particle.last_calculated_fitness = particle.fitness

            if particle.fitness >= particle.pBestFitness and particle.fitness != 0:
                particle.pBest = particle.position.copy()
                particle.pBestFitness = particle.fitness

            # 更新全局最优
            if particle.fitness >= self.gBestFitness and particle.fitness != 0:
                self.gBest = copy.deepcopy(particle)
                self.gBestFitness = particle.fitness

            # 更新当前代最优
            if particle.fitness > self.currentGenBestFitness:
                self.currentGenBest = copy.deepcopy(particle)
                self.currentGenBestFitness = particle.fitness

        # 4. 【周期性发送】定期向随机配对伙伴发送当前代最优位置
        if (self.partner_rank is not None and
                self.global_comm is not None and
                not self.global_solution_found and
                not self.partner_finished and
                self.iteration % PSO.COMM_FREQUENCY == 0):
            start_comm = time.time()
            req = self.global_comm.isend(self.currentGenBest.position, dest=self.partner_rank, tag=100)
            req.Wait()
            end_comm = time.time()
            self.comm_time += (end_comm - start_comm)

    def calculateFitnessBasedOnPath(self, position):
        self.evaluation_count += 1  # 🔧 新增:每次评价时计数
        x, y, z = position[0], position[1], position[2]
        if self.mode == "full":
            generated_path = generate_path_full(x, y, z)
        elif self.mode == "partial":
            generated_path = generate_path_partial(x, y, z)
        elif self.mode == "path3":
            generated_path = generate_path3(x, y, z)
        else:
            generated_path = []

        target_counter = Counter(self.target_path)
        candidate_counter = Counter(generated_path)
        matched = 0
        for code, req_count in target_counter.items():
            matched += min(candidate_counter.get(code, 0), req_count)
        return matched / len(self.target_path) if len(self.target_path) > 0 else 0.0

    def getTopParticles(self, top_n=2):
        sorted_particles = sorted(self.swarm, key=lambda p: p.fitness, reverse=True)
        return sorted_particles[:top_n]

    def incorporateFeedbackParticles(self, global_particles):
        """
        从全局优化中获取反馈粒子，并随机替换一些本地粒子
        global_particles: 来自进程0的全局优秀粒子列表，每个元素是一个形如[x, y, z]的位置
        """
        if not global_particles or len(global_particles) == 0:
            return

        # 决定要替换的粒子索引
        num_to_replace = min(len(global_particles), len(self.swarm) // 3)
        replace_indices = random.sample(range(len(self.swarm)), num_to_replace)

        for i, idx in enumerate(replace_indices):
            if i >= len(global_particles):
                break

            # 创建新粒子
            new_position = global_particles[i]

            # 保存旧速度方向，但减小幅度以允许探索
            old_velocity = self.swarm[idx].velocity.copy()
            new_velocity = [v * 0.5 for v in old_velocity]

            # 更新粒子
            self.swarm[idx].position = new_position
            self.swarm[idx].velocity = new_velocity

            # 计算新适应度
            new_fitness = self.calculateFitnessBasedOnPath(new_position)
            self.swarm[idx].fitness = new_fitness
            self.swarm[idx].last_calculated_fitness = new_fitness

            # 更新个体最优
            if new_fitness >= self.swarm[idx].pBestFitness and new_fitness != 0:
                self.swarm[idx].pBest = new_position.copy()
                self.swarm[idx].pBestFitness = new_fitness

            # 更新全局最优
            if new_fitness >= self.gBestFitness and new_fitness != 0:
                self.gBest = copy.deepcopy(self.swarm[idx])
                self.gBestFitness = new_fitness


# --- 全局5D PSO类（用于进程0，全局优化） ---
class PSO5D:
    DIMENSION = 5
    SWARM_SIZE = 26
    MAX_ITERATIONS = 8000
    MIN_SPEED = -100
    MAX_SPEED = 100
    W = 0.729
    C1 = 2.0
    C2 = 2.0
    COMM_FREQUENCY = 5  # 全局粒子通信频率

    class Particle:
        def __init__(self, dimension):
            self.position = [0] * dimension  # 五维位置
            self.velocity = [0] * dimension  # 五维速度
            self.fitness = float('-inf')  # 当前适应度
            self.pBest = [0] * dimension  # 个人历史最优位置
            self.pBestFitness = float('-inf')  # 个人历史最优适应度
            self.last_calculated_fitness = None
            self.paths = None  # 存储通信获取的子路径

    def __init__(self, targetPaths, comm, global_comm=None, partner_rank=None, group_color=0):
        # targetPaths 包含四个子目标路径
        self.targetPaths = targetPaths
        self.comm = comm
        self.swarm = [PSO5D.Particle(PSO5D.DIMENSION) for _ in range(PSO5D.SWARM_SIZE)]
        self.gBest = PSO5D.Particle(PSO5D.DIMENSION)
        self.gBestFitness = float('-inf')

        # 添加当前代最优粒子属性
        self.currentGenBest = PSO5D.Particle(PSO5D.DIMENSION)  # 当前代最优粒子
        self.currentGenBestFitness = float('-inf')  # 当前代最优适应度

        self.rand = random.Random()
        self.iteration = 0  # 迭代计数器
        self.evaluation_count = 0  # 🔧 新增:评价次数计数器

        # 随机配对通信相关属性
        self.global_comm = global_comm  # 全局MPI通信器
        self.partner_rank = partner_rank  # 随机配对伙伴进程的rank
        self.partner_gBest = None  # 伙伴的全局最优粒子
        self.partner_fitness = float('-inf')  # 伙伴的最优适应度

        self.comm_time = 0.0  # 通信时间计数器
        self.partner_finished = False  # 标志随机配对伙伴是否已完成
        self.group_color = group_color  # 组颜色/ID，用于日志

        # 初始化粒子群
        self.initializeSwarm()

    def initializeSwarm(self):
        for particle in self.swarm:
            for d in range(PSO5D.DIMENSION):
                # 使用对应维度的物理范围
                rangeMin, rangeMax = POSITION_RANGES[d]
                particle.position[d] = int(round(rangeMin + self.rand.random() * (rangeMax - rangeMin)))
                # 速度初始化使用相对范围
                velocity_range = (rangeMax - rangeMin) / 20  # 使用范围的1/20作为速度范围
                particle.velocity[d] = int(round(-velocity_range + self.rand.random() * 2 * velocity_range))
                particle.pBest[d] = particle.position[d]

    def notifyPartnerOfGlobalTermination(self):
        """当找到最优解时，通知伙伴进程"""
        if self.partner_rank is not None and self.global_comm is not None and not self.partner_finished:
            start_comm = time.time()
            req = self.global_comm.isend("finished", dest=self.partner_rank, tag=300)
            req.Wait()
            end_comm = time.time()
            self.comm_time += (end_comm - start_comm)
            self.partner_finished = True  # 设置为已通知

    def initializeSwarmFromLocalPSO(self, p1_particles, p2_particles, p3_particles):
        top_global_particles = sorted(self.swarm, key=lambda p: p.fitness, reverse=True)[:2]
        new_swarm = []
        for i in range(min(2, len(top_global_particles))):
            new_swarm.append(copy.deepcopy(top_global_particles[i]))

        k = len(new_swarm)  # Start counting from where we left off

        # 使用三重组合方式
        for i in range(len(p1_particles)):
            for j in range(len(p2_particles)):
                if k < PSO5D.SWARM_SIZE:
                    for m in range(len(p3_particles)):
                        # 组合模式1：P1优先，保留xyz
                        if k < PSO5D.SWARM_SIZE:
                            new_particle1 = PSO5D.Particle(PSO5D.DIMENSION)
                            new_particle1.position = [
                                p1_particles[i][0],  # P1的x
                                p1_particles[i][1],  # P1的y
                                p1_particles[i][2],  # P1的z
                                p2_particles[j][2],  # P2的w
                                p3_particles[m][2]  # P3的m
                            ]

                            # 确保位置在有效范围内
                            for dim in range(PSO5D.DIMENSION):
                                rangeMin, rangeMax = POSITION_RANGES[dim]
                                new_particle1.position[dim] = int(
                                    round(max(rangeMin, min(rangeMax, new_particle1.position[dim]))))
                                # 初始化速度
                                velocity_range = (rangeMax - rangeMin) / 20
                                new_particle1.velocity[dim] = int(
                                    round(-velocity_range + self.rand.random() * 2 * velocity_range))

                            new_particle1.pBest = new_particle1.position.copy()
                            new_swarm.append(new_particle1)
                            k += 1

                        # 组合模式2：混合维度
                        if k < PSO5D.SWARM_SIZE:
                            new_particle2 = PSO5D.Particle(PSO5D.DIMENSION)
                            new_particle2.position = [
                                p1_particles[i][0],  # x from P1 ✓
                                p2_particles[j][0],  # y from P2 (P2局部[0]=全局y) ✓
                                p2_particles[j][1],  # z from P2 (P2局部[1]=全局z) ✓
                                p2_particles[j][2],  # w from P2 ✓
                                p3_particles[m][2]  # m from P3 ✓
                            ]

                            # 确保位置在有效范围内
                            for dim in range(PSO5D.DIMENSION):
                                rangeMin, rangeMax = POSITION_RANGES[dim]
                                new_particle2.position[dim] = int(
                                    round(max(rangeMin, min(rangeMax, new_particle2.position[dim]))))
                                velocity_range = (rangeMax - rangeMin) / 20
                                new_particle2.velocity[dim] = int(
                                    round(-velocity_range + self.rand.random() * 2 * velocity_range))

                            new_particle2.pBest = new_particle2.position.copy()
                            new_swarm.append(new_particle2)
                            k += 1

                        # 组合模式3：P3优先
                        if k < PSO5D.SWARM_SIZE:
                            new_particle3 = PSO5D.Particle(PSO5D.DIMENSION)
                            new_particle3.position = [
                                p1_particles[i][0],  # x from P1 ✓
                                p2_particles[j][0],  # y from P2 (P2局部[0]=全局y) ✓
                                p3_particles[m][0],  # z from P3 (P3局部[0]=全局z) ✓
                                p3_particles[m][1],  # w from P3 (P3局部[1]=全局w) ✓
                                p3_particles[m][2]  # m from P3 ✓
                            ]

                            # 确保位置在有效范围内
                            for dim in range(PSO5D.DIMENSION):
                                rangeMin, rangeMax = POSITION_RANGES[dim]
                                new_particle3.position[dim] = int(
                                    round(max(rangeMin, min(rangeMax, new_particle3.position[dim]))))
                                velocity_range = (rangeMax - rangeMin) / 20
                                new_particle3.velocity[dim] = int(
                                    round(-velocity_range + self.rand.random() * 2 * velocity_range))

                            new_particle3.pBest = new_particle3.position.copy()
                            new_swarm.append(new_particle3)
                            k += 1

                        if k >= PSO5D.SWARM_SIZE:
                            break
                    if k >= PSO5D.SWARM_SIZE:
                        break
                if k >= PSO5D.SWARM_SIZE:
                    break

        self.swarm = new_swarm

        # 更新全局最优
        self.gBestFitness = float('-inf')
        for particle in self.swarm:
            if particle.fitness >= self.gBestFitness and particle.fitness is not None and particle.fitness != 0:
                self.gBestFitness = particle.fitness
                self.gBest = copy.deepcopy(particle)

        return len(new_swarm)

    def getTopParticlesForFeedback(self, top_n=H):
        """获取性能最佳的top_n个粒子用于反馈给局部优化进程"""
        sorted_particles = sorted(self.swarm, key=lambda p: p.fitness, reverse=True)
        top_particles = sorted_particles[:top_n]

        # 为三个进程准备反馈，修改为3维反馈
        p1_feedback = [[p.position[0], p.position[1], p.position[2]] for p in top_particles]  # xyz
        p2_feedback = [[p.position[1], p.position[2], p.position[3]] for p in top_particles]  # yzw
        p3_feedback = [[p.position[2], p.position[3], p.position[4]] for p in top_particles]  # zwm

        return p1_feedback, p2_feedback, p3_feedback

    def generateMPIPaths(self, a):
        comm = self.comm
        status = MPI.Status()
        path = []
        x = a[0]
        y = a[1]
        z = a[2]
        w = a[3]
        m = a[4]

        # 发送3个参数给每个工作进程
        # 给进程1发送x,y,z
        comm.send(x, dest=1, tag=1)
        comm.send(y, dest=1, tag=2)
        comm.send(z, dest=1, tag=3)

        # 给进程2发送y,z,w
        comm.send(y, dest=2, tag=1)
        comm.send(z, dest=2, tag=2)
        comm.send(w, dest=2, tag=3)

        # 给进程3发送z,w,m
        comm.send(z, dest=3, tag=1)
        comm.send(w, dest=3, tag=2)
        comm.send(m, dest=3, tag=3)
        if (x ** 0.5 + y > z * w - 200) != ((x + 20000) ** 0.5 + y > z * w - 200):
            path.append(1)
        if (x ** 0.5 + y > z * w - 200) != (x ** 2 + y > z * w - 200):
            path.append(2)
        if (x ** 0.5 + y > z * w - 200) != (x ** 0.5 + y * 5 > z * w - 200):
            path.append(3)
        if (x ** 0.5 + y > z * w - 200) != (x ** 0.5 + y + 6000 > z * w - 200):
            path.append(4)
        if (x ** 0.5 + y > z * w - 200) != (x ** 0.5 + y > z * w / 5 - 200):
            path.append(5)
        if (x ** 0.5 + y > z * w - 200) != (x ** 0.5 + y > z * w - 20000):
            path.append(6)
        if (x ** 0.5 + y > z * w - 200) != (x ** 0.5 + y > w / 8 * w - 200):
            path.append(7)
        if (x ** 0.5 + y > z * w - 200) != (x ** 0.5 + y > z * z / 15 - 200):
            path.append(8)
        if (x ** 0.5 + y > z * w - 200) != (x ** 0.5 + y > z * x / 80 - 200):
            path.append(9)
        if (x ** 0.5 + y > z * w - 200) != (x ** 0.5 + y > x / 1000 * w - 200):
            path.append(10)

        if x ** 0.5 + y > z * w - 200:
            type_code = 'A1'

        if (m ** (y // 20) < 50) != ((m * 5) ** (y // 20) < 50):
            path.append(11)
        if (m ** (y // 20) < 50) != ((m + 10) ** (y // 20) < 50):
            path.append(12)
        if (m ** (y // 20) < 50) != (m ** (y + 20) < 50):
            path.append(13)
        if (m ** (y // 20) < 50) != (m ** (y / 20) < 50):
            path.append(14)
        if (m ** (y // 20) < 50) != (m ** (y * 20) < 50):
            path.append(15)
        if (m ** (y // 20) < 50) != (m ** (y * 2 // 20) < 50):
            path.append(16)
        if (m ** (y // 20) < 50) != (m ** ((y + 50) // 20) < 50):
            path.append(17)
        if (m ** (y // 20) < 50) != (m ** (y // 10) < 50):
            path.append(18)
        if (m ** (y // 20) < 50) != (m ** (y // 20) < 150):
            path.append(19)
        if (m ** (y // 20) < 50) != (m ** (y // 20) < 50 + 10 * m):
            path.append(20)

        if m ** (y // 20) < 50:
            type_code = 'A2'

        if (x // 100 + z * m > w + y) != (x * 2 // 100 + z * m > w + y):
            path.append(21)
        if (x // 100 + z * m > w + y) != ((x + 2500) // 100 + z * m > w + y):
            path.append(22)
        if (x // 100 + z * m > w + y) != (x // 50 + z * m > w + y):
            path.append(23)
        if (x // 100 + z * m > w + y) != (x // 100 + z * 2 * m > w + y):
            path.append(24)
        if (x // 100 + z * m > w + y) != (x // 100 + z * z > w + y):
            path.append(25)
        if (x // 100 + z * m > w + y) != (x // 100 + m * m > w + y):
            path.append(26)
        if (x // 100 + z * m > w + y) != (x // 100 + z * m > w / 15 + y):
            path.append(27)
        if (x // 100 + z * m > w + y) != (x // 100 + z * m > w + y / 2):
            path.append(28)
        if (x // 100 + z * m > w + y) != (x // 100 + z * m > (w + y) / 1.5):
            path.append(29)
        if (x // 100 + z * m > w + y) != (x // 100 + z * m > w + y - 20):
            path.append(30)

        if x // 100 + z * m > w + y:
            type_code = 'A3'

        # 接收来自工作进程的结果
        path1 = comm.recv(source=1, tag=11, status=status)
        path2 = comm.recv(source=2, tag=21, status=status)
        path3 = comm.recv(source=3, tag=31, status=status)

        # 主进程基于 x,y,z,w,m 进行判断，生成覆盖码

        return [path, path1, path2, path3]

    def calculateFitnessBasedOnPaths(self, mpi_paths, targetPaths):
        total_fitness = 0.0
        for group_index in range(4):  # 修正：仍然是4条路径
            target = targetPaths[group_index]
            candidate = mpi_paths[group_index] if group_index < len(mpi_paths) else []
            target_counter = Counter(target)
            candidate_counter = Counter(candidate)
            matched = 0
            for code, req_count in target_counter.items():
                matched += min(candidate_counter.get(code, 0), req_count)
            sub_fitness = matched / len(target) if len(target) > 0 else 0.0
            total_fitness += sub_fitness
        return total_fitness / 4.0  # 修正：仍然除以4

    def updateParticles(self, currentTargetPaths):
        self.iteration += 1  # 增加迭代计数

        # 重置当前代最优粒子
        self.currentGenBestFitness = float('-inf')
        self.currentGenBest = PSO5D.Particle(PSO5D.DIMENSION)

        # 1. 【配对接收】检查来自随机配对伙伴的消息（非阻塞）
        if self.partner_rank is not None and self.global_comm is not None and not self.partner_finished:
            while self.global_comm.Iprobe(source=self.partner_rank, tag=300):
                start_comm = time.time()
                partner_data = self.global_comm.recv(source=self.partner_rank, tag=300)
                end_comm = time.time()
                self.comm_time += (end_comm - start_comm)

                if partner_data == "finished":
                    self.partner_finished = True
                    self.partner_gBest = None
                    self.partner_fitness = float('-inf')
                else:
                    partner_position = partner_data['position']
                    partner_reported_fitness = partner_data['fitness']
                    partner_paths = partner_data['paths']

                    local_fitness = self.calculateFitnessBasedOnPaths(partner_paths, self.targetPaths)

                    if local_fitness == OPTIMAL_FITNESS:
                        self.gBest = PSO5D.Particle(PSO5D.DIMENSION)
                        self.gBest.position = partner_position.copy()
                        self.gBest.paths = partner_paths.copy()
                        self.gBestFitness = OPTIMAL_FITNESS
                        print(f"[全局进程，第{self.group_color}组] 在评估高相似度伙伴粒子时找到最优解！")
                        print(f"[全局进程，第{self.group_color}组] 最优粒子位置: {partner_position}")
                        break

                    elif local_fitness >= self.partner_fitness and local_fitness != 0:
                        self.partner_gBest = PSO5D.Particle(PSO5D.DIMENSION)
                        self.partner_gBest.position = partner_position.copy()
                        self.partner_gBest.paths = partner_paths.copy()
                        self.partner_fitness = local_fitness

        # 2. 确定是否使用配对伙伴信息
        use_partner = False
        if self.partner_gBest is not None:
            # 直接进行适应度判断，去掉概率性接受
            if self.gBestFitness == float('-inf'):
                use_partner = True
            else:
                if self.partner_fitness > self.gBestFitness or abs(self.gBestFitness - self.partner_fitness) < 0.3:
                    use_partner = True

        # 3. 【核心计算】更新粒子
        for particle in self.swarm:
            for i in range(PSO5D.DIMENSION):
                r1 = self.rand.random()
                r2 = self.rand.random()

                if use_partner:
                    r3 = self.rand.random()
                    particle.velocity[i] = (self.W * particle.velocity[i] +
                                            self.C1 * r1 * (particle.pBest[i] - particle.position[i]) +
                                            self.C2 * r2 * (self.gBest.position[i] - particle.position[i]) +
                                            self.C1 * r3 * (self.partner_gBest.position[i] - particle.position[i]))
                else:
                    particle.velocity[i] = (self.W * particle.velocity[i] +
                                            self.C1 * r1 * (particle.pBest[i] - particle.position[i]) +
                                            self.C2 * r2 * (self.gBest.position[i] - particle.position[i]))

                particle.velocity[i] = max(PSO5D.MIN_SPEED, min(PSO5D.MAX_SPEED, particle.velocity[i]))
                particle.position[i] += int(round(particle.velocity[i]))

                # 使用对应维度的边界约束
                rangeMin, rangeMax = POSITION_RANGES[i]
                particle.position[i] = int(round(max(rangeMin, min(rangeMax, particle.position[i]))))

            mpi_result = self.generateMPIPaths(particle.position)
            particle.paths = mpi_result
            fitness = self.calculateFitnessBasedOnPaths(mpi_result, currentTargetPaths)
            particle.fitness = fitness
            self.evaluation_count += 1  # 🔧 新增:每次评价时计数

            if fitness >= particle.pBestFitness and fitness != 0:
                particle.pBest = particle.position.copy()
                particle.pBestFitness = fitness

            if fitness >= self.gBestFitness and fitness != 0:
                self.gBest = copy.deepcopy(particle)
                self.gBestFitness = fitness

            if fitness > self.currentGenBestFitness:
                self.currentGenBest = copy.deepcopy(particle)
                self.currentGenBestFitness = fitness

        # 4. 【周期性发送】定期向随机配对伙伴发送当前代最优位置
        if (self.partner_rank is not None and
                self.global_comm is not None and
                self.gBestFitness < OPTIMAL_FITNESS and
                not self.partner_finished and
                self.iteration % self.COMM_FREQUENCY == 0):
            start_comm = time.time()
            data_to_send = {
                'position': self.currentGenBest.position,
                'fitness': self.currentGenBestFitness,
                'paths': self.currentGenBest.paths
            }
            req = self.global_comm.isend(data_to_send, dest=self.partner_rank, tag=300)
            req.Wait()
            end_comm = time.time()
            self.comm_time += (end_comm - start_comm)


# --- 修改后的路径生成函数 ---
def generate_path_full(x, y, z):
    path = []
    if ((x * y) ** (z // 4) > 500000) != ((x * y * 5) ** (z // 4) > 500000):
        path.append(1)
    if ((x * y) ** (z // 4) > 500000) != ((x * (y + 150)) ** (z // 4) > 500000):
        path.append(2)
    if ((x * y) ** (z // 4) > 500000) != (((x + 6000) * y) ** (z // 4) > 500000):
        path.append(3)
    if ((x * y) ** (z // 4) > 500000) != ((x * y) ** (z * 2 // 4) > 500000):
        path.append(4)
    if ((x * y) ** (z // 4) > 500000) != ((x * y) ** (z / 4) > 500000):
        path.append(5)
    if ((x * y) ** (z // 4) > 500000) != ((x * y) ** (z // 2) > 500000):
        path.append(6)
    if ((x * y) ** (z // 4) > 500000) != ((x * y) ** (z // 4) > 5000):
        path.append(7)
    if ((x * y) ** (z // 4) > 500000) != ((x * y) ** (z // 4) > 500000 - 250 * x):
        path.append(8)
    if ((x * y) ** (z // 4) > 500000) != ((x * y) * (z // 4) > 500000):
        path.append(9)
    if ((x * y) ** (z // 4) > 500000) != ((x * y) ** (z // 4) > 500000 / (5 * y + 800)):
        path.append(10)

    if (x * y) ** (z // 4) > 500000:
        type_code = 'B1'

    if (x % (y + z) > 100) != (x % (y + z) > y / 1.2):
        path.append(11)
    if (x % (y + z) > 100) != (x % (y + z) > x / 20):
        path.append(12)
    if (x % (y + z) > 100) != (x % (y + z) > 100 / (x / 8050 + 2)):
        path.append(13)
    if (x % (y + z) > 100) != (x % (y + z) > 100 - x / 68):
        path.append(14)
    if (x % (y + z) > 100) != (x % (y + z) > 50):
        path.append(15)
    if (x % (y + z) > 100) != (x % (y * 2 + z) > 100):
        path.append(16)
    if (x % (y + z) > 100) != (x % (y + z * 10) > 100):
        path.append(17)
    if (x % (y + z) > 100) != (x % (y + z + 50) > 100):
        path.append(18)
    if (x % (y + z) > 100) != (x % (y + y * 1.2) > 100):
        path.append(19)
    if (x % (y + z) > 100) != (x % (z * 15 + z) > 100):
        path.append(20)

    if x % (y + z) > 100:
        type_code = 'B2'

    if (x ** 2 - y * z > 8000) != ((x / 15) ** 2 - y * z > 8000):
        path.append(21)
    if (x ** 2 - y * z > 8000) != (x ** 1.4 - y * z > 8000):
        path.append(22)
    if (x ** 2 - y * z > 8000) != (x * (x - 1200) - y * z > 8000):
        path.append(23)
    if (x ** 2 - y * z > 8000) != (x ** 2 - y * z * 1210 > 8000):
        path.append(24)
    if (x ** 2 - y * z > 8000) != (x ** 2 - y * (z + 7050) > 8000):
        path.append(25)
    if (x ** 2 - y * z > 8000) != (x ** 2 - y * 8000 > 8000):
        path.append(26)
    if (x ** 2 - y * z > 8000) != (x ** 2 - y * y * 25 * z > 8000):
        path.append(27)
    if (x ** 2 - y * z > 8000) != (x ** 2 - y * z * z * 200 > 8000):
        path.append(28)
    if (x ** 2 - y * z > 8000) != (x ** 2 - y * z * x * 2 > 8000):
        path.append(29)
    if (x ** 2 - y * z > 8000) != (x ** 2 - y * x * 15 > 8000):
        path.append(30)

    if x ** 2 - y * z > 8000:
        type_code = 'B3'

    return path


def mutation_process1(first_val, second_val, third_val):
    x = first_val
    y = second_val
    z = third_val
    path = []
    if ((x * y) ** (z // 4) > 500000) != ((x * y * 5) ** (z // 4) > 500000):
        path.append(1)
    if ((x * y) ** (z // 4) > 500000) != ((x * (y + 150)) ** (z // 4) > 500000):
        path.append(2)
    if ((x * y) ** (z // 4) > 500000) != (((x + 6000) * y) ** (z // 4) > 500000):
        path.append(3)
    if ((x * y) ** (z // 4) > 500000) != ((x * y) ** (z * 2 // 4) > 500000):
        path.append(4)
    if ((x * y) ** (z // 4) > 500000) != ((x * y) ** (z / 4) > 500000):
        path.append(5)
    if ((x * y) ** (z // 4) > 500000) != ((x * y) ** (z // 2) > 500000):
        path.append(6)
    if ((x * y) ** (z // 4) > 500000) != ((x * y) ** (z // 4) > 5000):
        path.append(7)
    if ((x * y) ** (z // 4) > 500000) != ((x * y) ** (z // 4) > 500000 - 250 * x):
        path.append(8)
    if ((x * y) ** (z // 4) > 500000) != ((x * y) * (z // 4) > 500000):
        path.append(9)
    if ((x * y) ** (z // 4) > 500000) != ((x * y) ** (z // 4) > 500000 / (5 * y + 800)):
        path.append(10)

    if (x * y) ** (z // 4) > 500000:
        type_code = 'B1'

    if (x % (y + z) > 100) != (x % (y + z) > y / 1.2):
        path.append(11)
    if (x % (y + z) > 100) != (x % (y + z) > x / 20):
        path.append(12)
    if (x % (y + z) > 100) != (x % (y + z) > 100 / (x / 8050 + 2)):
        path.append(13)
    if (x % (y + z) > 100) != (x % (y + z) > 100 - x / 68):
        path.append(14)
    if (x % (y + z) > 100) != (x % (y + z) > 50):
        path.append(15)
    if (x % (y + z) > 100) != (x % (y * 2 + z) > 100):
        path.append(16)
    if (x % (y + z) > 100) != (x % (y + z * 10) > 100):
        path.append(17)
    if (x % (y + z) > 100) != (x % (y + z + 50) > 100):
        path.append(18)
    if (x % (y + z) > 100) != (x % (y + y * 1.2) > 100):
        path.append(19)
    if (x % (y + z) > 100) != (x % (z * 15 + z) > 100):
        path.append(20)

    if x % (y + z) > 100:
        type_code = 'B2'

    if (x ** 2 - y * z > 8000) != ((x / 15) ** 2 - y * z > 8000):
        path.append(21)
    if (x ** 2 - y * z > 8000) != (x ** 1.4 - y * z > 8000):
        path.append(22)
    if (x ** 2 - y * z > 8000) != (x * (x - 1200) - y * z > 8000):
        path.append(23)
    if (x ** 2 - y * z > 8000) != (x ** 2 - y * z * 1210 > 8000):
        path.append(24)
    if (x ** 2 - y * z > 8000) != (x ** 2 - y * (z + 7050) > 8000):
        path.append(25)
    if (x ** 2 - y * z > 8000) != (x ** 2 - y * 8000 > 8000):
        path.append(26)
    if (x ** 2 - y * z > 8000) != (x ** 2 - y * y * 25 * z > 8000):
        path.append(27)
    if (x ** 2 - y * z > 8000) != (x ** 2 - y * z * z * 200 > 8000):
        path.append(28)
    if (x ** 2 - y * z > 8000) != (x ** 2 - y * z * x * 2 > 8000):
        path.append(29)
    if (x ** 2 - y * z > 8000) != (x ** 2 - y * x * 15 > 8000):
        path.append(30)

    if x ** 2 - y * z > 8000:
        type_code = 'B3'

    return path


def generate_path_partial(y, z, w):
    path = []
    if (y ** (w // 15) < z * 20) != ((y * 5) ** (w // 15) < z * 20):
        path.append(1)
    if (y ** (w // 15) < z * 20) != ((y + 200) ** (w // 15) < z * 20):
        path.append(2)
    if (y ** (w // 15) < z * 20) != (3000 ** (w // 15) < z * 20):
        path.append(3)
    if (y ** (w // 15) < z * 20) != (y ** (w + 15) < z * 20):
        path.append(4)
    if (y ** (w // 15) < z * 20) != (y ** (w / 15) < z * 20):
        path.append(5)
    if (y ** (w // 15) < z * 20) != (y ** (w * 15) < z * 20):
        path.append(6)
    if (y ** (w // 15) < z * 20) != (y ** (w // 2) < z * 20):
        path.append(7)
    if (y ** (w // 15) < z * 20) != (y ** (w * 5 // 15) < z * 20):
        path.append(8)
    if (y ** (w // 15) < z * 20) != (y ** ((w + 20) // 15) < z * 20):
        path.append(9)
    if (y ** (w // 15) < z * 20) != (y ** (w // 15) < z * 4):
        path.append(10)

    if y ** (w // 15) < z * 20:
        type_code = 'C1'

    if ((y + w) ** 0.5 > z + 30) != ((y + w * 25) ** 0.5 > z + 30):
        path.append(11)
    if ((y + w) ** 0.5 > z + 30) != ((y * 15 + w) ** 0.5 > z + 30):
        path.append(12)
    if ((y + w) ** 0.5 > z + 30) != ((y + y * 15) ** 0.5 > z + 30):
        path.append(13)
    if ((y + w) ** 0.5 > z + 30) != ((w * 25 + w) ** 0.5 > z + 30):
        path.append(14)
    if ((y + w) ** 0.5 > z + 30) != ((y + 1000) ** 0.5 > z + 30):
        path.append(15)
    if ((y + w) ** 0.5 > z + 30) != ((1000 + w) ** 0.5 > z + 30):
        path.append(16)
    if ((y + w) ** 0.5 > z + 30) != ((y + z * 150) ** 0.5 > z + 30):
        path.append(17)
    if ((y + w) ** 0.5 > z + 30) != ((z * 155 + w) ** 0.5 > z + 30):
        path.append(18)
    if ((y + w) ** 0.5 > z + 30) != ((y + w + 150 * z) ** 0.5 > z + 30):
        path.append(19)
    if ((y + w) ** 0.5 > z + 30) != ((y + w + 15 * y) ** 0.5 > z + 30):
        path.append(20)

    if (y + w) ** 0.5 > z + 30:
        type_code = 'C2'

    if (y * w // 10 > z * 3) != (y * w / 5 // 10 > z * 3):
        path.append(21)
    if (y * w // 10 > z * 3) != (y * (w - 15) // 10 > z * 3):
        path.append(22)
    if (y * w // 10 > z * 3) != ((y - 35) * w // 10 > z * 3):
        path.append(23)
    if (y * w // 10 > z * 3) != (y * w // 50 > z * 3):
        path.append(24)
    if (y * w // 10 > z * 3) != (y * w // 10 > z * 3):
        path.append(25)
    if (y * w // 10 > z * 3) != (y * 5 // 10 > z * 3):
        path.append(26)
    if (y * w // 10 > z * 3) != (10 * w // 10 > z * 3):
        path.append(27)
    if (y * w // 10 > z * 3) != (y * z / 1.5 // 10 > z * 3):
        path.append(28)
    if (y * w // 10 > z * 3) != (z / 1.08 * w // 10 > z * 3):
        path.append(29)
    if (y * w // 10 > z * 3) != (y * w // 10 > (z + 30) * 3):
        path.append(30)

    if y * w // 10 > z * 3:
        type_code = 'C3'

    return path


def mutation_process2(first_val, second_val, third_val):
    y = first_val
    z = second_val
    w = third_val
    path = []
    if (y ** (w // 15) < z * 20) != ((y * 5) ** (w // 15) < z * 20):
        path.append(1)
    if (y ** (w // 15) < z * 20) != ((y + 200) ** (w // 15) < z * 20):
        path.append(2)
    if (y ** (w // 15) < z * 20) != (3000 ** (w // 15) < z * 20):
        path.append(3)
    if (y ** (w // 15) < z * 20) != (y ** (w + 15) < z * 20):
        path.append(4)
    if (y ** (w // 15) < z * 20) != (y ** (w / 15) < z * 20):
        path.append(5)
    if (y ** (w // 15) < z * 20) != (y ** (w * 15) < z * 20):
        path.append(6)
    if (y ** (w // 15) < z * 20) != (y ** (w // 2) < z * 20):
        path.append(7)
    if (y ** (w // 15) < z * 20) != (y ** (w * 5 // 15) < z * 20):
        path.append(8)
    if (y ** (w // 15) < z * 20) != (y ** ((w + 20) // 15) < z * 20):
        path.append(9)
    if (y ** (w // 15) < z * 20) != (y ** (w // 15) < z * 4):
        path.append(10)

    if y ** (w // 15) < z * 20:
        type_code = 'C1'

    if ((y + w) ** 0.5 > z + 30) != ((y + w * 25) ** 0.5 > z + 30):
        path.append(11)
    if ((y + w) ** 0.5 > z + 30) != ((y * 15 + w) ** 0.5 > z + 30):
        path.append(12)
    if ((y + w) ** 0.5 > z + 30) != ((y + y * 15) ** 0.5 > z + 30):
        path.append(13)
    if ((y + w) ** 0.5 > z + 30) != ((w * 25 + w) ** 0.5 > z + 30):
        path.append(14)
    if ((y + w) ** 0.5 > z + 30) != ((y + 1000) ** 0.5 > z + 30):
        path.append(15)
    if ((y + w) ** 0.5 > z + 30) != ((1000 + w) ** 0.5 > z + 30):
        path.append(16)
    if ((y + w) ** 0.5 > z + 30) != ((y + z * 150) ** 0.5 > z + 30):
        path.append(17)
    if ((y + w) ** 0.5 > z + 30) != ((z * 155 + w) ** 0.5 > z + 30):
        path.append(18)
    if ((y + w) ** 0.5 > z + 30) != ((y + w + 150 * z) ** 0.5 > z + 30):
        path.append(19)
    if ((y + w) ** 0.5 > z + 30) != ((y + w + 15 * y) ** 0.5 > z + 30):
        path.append(20)

    if (y + w) ** 0.5 > z + 30:
        type_code = 'C2'

    if (y * w // 10 > z * 3) != (y * w / 5 // 10 > z * 3):
        path.append(21)
    if (y * w // 10 > z * 3) != (y * (w - 15) // 10 > z * 3):
        path.append(22)
    if (y * w // 10 > z * 3) != ((y - 35) * w // 10 > z * 3):
        path.append(23)
    if (y * w // 10 > z * 3) != (y * w // 50 > z * 3):
        path.append(24)
    if (y * w // 10 > z * 3) != (y * w // 10 > z * 3):
        path.append(25)
    if (y * w // 10 > z * 3) != (y * 5 // 10 > z * 3):
        path.append(26)
    if (y * w // 10 > z * 3) != (10 * w // 10 > z * 3):
        path.append(27)
    if (y * w // 10 > z * 3) != (y * z / 1.5 // 10 > z * 3):
        path.append(28)
    if (y * w // 10 > z * 3) != (z / 1.08 * w // 10 > z * 3):
        path.append(29)
    if (y * w // 10 > z * 3) != (y * w // 10 > (z + 30) * 3):
        path.append(30)

    if y * w // 10 > z * 3:
        type_code = 'C3'

    return path


def generate_path3(z, w, m):
    path = []
    if (z ** (m // 2) > w * 100) != ((z + 20) ** (m // 2) > w * 100):
        path.append(1)
    if (z ** (m // 2) > w * 100) != ((z * 2) ** (m // 2) > w * 100):
        path.append(2)
    if (z ** (m // 2) > w * 100) != (50 ** (m // 2) > w * 100):
        path.append(3)
    if (z ** (m // 2) > w * 100) != (z ** (m * 2 // 2) > w * 100):
        path.append(4)
    if (z ** (m // 2) > w * 100) != (z ** (2 // 2) > w * 100):
        path.append(5)
    if (z ** (m // 2) > w * 100) != (z ** (m // 2) > w * 10):
        path.append(6)
    if (z ** (m // 2) > w * 100) != (z ** (m // 2) > (w - 20) * 100):
        path.append(7)
    if (z ** (m // 2) > w * 100) != (z ** (m // 2) > 6 * 100):
        path.append(8)
    if (z ** (m // 2) > w * 100) != (z ** (m // 2) > (z - 2) * 100):
        path.append(9)
    if (z ** (m // 2) > w * 100) != (z ** (m // 2) > (m - 1.5) * 100):
        path.append(10)

    if z ** (m // 2) > w * 100:
        type_code = 'D1'

    if ((z * w) ** 0.5 + m > 85) != ((z * w * 15) ** 0.5 + m > 85):
        path.append(11)
    if ((z * w) ** 0.5 + m > 85) != ((z * (w + 700)) ** 0.5 + m > 85):
        path.append(12)
    if ((z * w) ** 0.5 + m > 85) != (((z + 100) * w) ** 0.5 + m > 85):
        path.append(13)
    if ((z * w) ** 0.5 + m > 85) != ((z * 600) ** 0.5 + m > 85):
        path.append(14)
    if ((z * w) ** 0.5 + m > 85) != ((100 * w) ** 0.5 + m > 85):
        path.append(15)
    if ((z * w) ** 0.5 + m > 85) != ((z * w) ** 0.5 + m > 25):
        path.append(16)
    if ((z * w) ** 0.5 + m > 85) != ((z * w) ** 0.5 + m * 8 > 85):
        path.append(17)
    if ((z * w) ** 0.5 + m > 85) != ((z * w) ** 0.5 + m > m * 6):
        path.append(18)
    if ((z * w) ** 0.5 + m > 85) != ((z * m * 150) ** 0.5 + m > 85):
        path.append(19)
    if ((z * w) ** 0.5 + m > 85) != ((z * w + 5500) ** 0.5 + m > 85):
        path.append(20)

    if (z * w) ** 0.5 + m > 85:
        type_code = 'D2'

    if (z * m % (w + 10) < 5) != (z * 5 * m % (w + 10) < 5):
        path.append(21)
    if (z * m % (w + 10) < 5) != ((z + 15) * m % (w + 10) < 5):
        path.append(22)
    if (z * m % (w + 10) < 5) != (z * (m + 15) % (w + 10) < 5):
        path.append(23)
    if (z * m % (w + 10) < 5) != (z * m % (w / 5 + 10) < 5):
        path.append(24)
    if (z * m % (w + 10) < 5) != (z * m % (w + 1) < 5):
        path.append(25)
    if (z * m % (w + 10) < 5) != (z * m % (w + 10) < 15):
        path.append(26)
    if (z * m % (w + 10) < 5) != (z * m % (5 + 10) < 5):
        path.append(27)
    if (z * m % (w + 10) < 5) != (z * z % (w + 10) < 5):
        path.append(28)
    if (z * m % (w + 10) < 5) != (m * 5 * m % (w + 10) < 5):
        path.append(29)
    if (z * m % (w + 10) < 5) != (z * 70 % (w + 10) < 5):
        path.append(30)

    if z * m % (w + 10) < 5:
        type_code = 'D3'

    return path


def mutation_process3(first_val, second_val, third_val):
    z = first_val
    w = second_val
    m = third_val
    path = []
    if (z ** (m // 2) > w * 100) != ((z + 20) ** (m // 2) > w * 100):
        path.append(1)
    if (z ** (m // 2) > w * 100) != ((z * 2) ** (m // 2) > w * 100):
        path.append(2)
    if (z ** (m // 2) > w * 100) != (50 ** (m // 2) > w * 100):
        path.append(3)
    if (z ** (m // 2) > w * 100) != (z ** (m * 2 // 2) > w * 100):
        path.append(4)
    if (z ** (m // 2) > w * 100) != (z ** (2 // 2) > w * 100):
        path.append(5)
    if (z ** (m // 2) > w * 100) != (z ** (m // 2) > w * 10):
        path.append(6)
    if (z ** (m // 2) > w * 100) != (z ** (m // 2) > (w - 20) * 100):
        path.append(7)
    if (z ** (m // 2) > w * 100) != (z ** (m // 2) > 6 * 100):
        path.append(8)
    if (z ** (m // 2) > w * 100) != (z ** (m // 2) > (z - 2) * 100):
        path.append(9)
    if (z ** (m // 2) > w * 100) != (z ** (m // 2) > (m - 1.5) * 100):
        path.append(10)

    if z ** (m // 2) > w * 100:
        type_code = 'D1'

    if ((z * w) ** 0.5 + m > 85) != ((z * w * 15) ** 0.5 + m > 85):
        path.append(11)
    if ((z * w) ** 0.5 + m > 85) != ((z * (w + 700)) ** 0.5 + m > 85):
        path.append(12)
    if ((z * w) ** 0.5 + m > 85) != (((z + 100) * w) ** 0.5 + m > 85):
        path.append(13)
    if ((z * w) ** 0.5 + m > 85) != ((z * 600) ** 0.5 + m > 85):
        path.append(14)
    if ((z * w) ** 0.5 + m > 85) != ((100 * w) ** 0.5 + m > 85):
        path.append(15)
    if ((z * w) ** 0.5 + m > 85) != ((z * w) ** 0.5 + m > 25):
        path.append(16)
    if ((z * w) ** 0.5 + m > 85) != ((z * w) ** 0.5 + m * 8 > 85):
        path.append(17)
    if ((z * w) ** 0.5 + m > 85) != ((z * w) ** 0.5 + m > m * 6):
        path.append(18)
    if ((z * w) ** 0.5 + m > 85) != ((z * m * 150) ** 0.5 + m > 85):
        path.append(19)
    if ((z * w) ** 0.5 + m > 85) != ((z * w + 5500) ** 0.5 + m > 85):
        path.append(20)

    if (z * w) ** 0.5 + m > 85:
        type_code = 'D2'

    if (z * m % (w + 10) < 5) != (z * 5 * m % (w + 10) < 5):
        path.append(21)
    if (z * m % (w + 10) < 5) != ((z + 15) * m % (w + 10) < 5):
        path.append(22)
    if (z * m % (w + 10) < 5) != (z * (m + 15) % (w + 10) < 5):
        path.append(23)
    if (z * m % (w + 10) < 5) != (z * m % (w / 5 + 10) < 5):
        path.append(24)
    if (z * m % (w + 10) < 5) != (z * m % (w + 1) < 5):
        path.append(25)
    if (z * m % (w + 10) < 5) != (z * m % (w + 10) < 15):
        path.append(26)
    if (z * m % (w + 10) < 5) != (z * m % (5 + 10) < 5):
        path.append(27)
    if (z * m % (w + 10) < 5) != (z * z % (w + 10) < 5):
        path.append(28)
    if (z * m % (w + 10) < 5) != (m * 5 * m % (w + 10) < 5):
        path.append(29)
    if (z * m % (w + 10) < 5) != (z * 70 % (w + 10) < 5):
        path.append(30)

    if z * m % (w + 10) < 5:
        type_code = 'D3'

    return path


def run_single_experiment(experiment_id, comm, rank, size, num_path_sets, procs_per_group):
    """运行单次实验"""
    # 设置随机种子
    set_experiment_seeds(experiment_id, rank)

    # 重置通信状态
    reset_communication_state(comm)

    # 确定进程是否参与计算
    if rank >= num_path_sets * procs_per_group:
        return None  # 不参与计算

    # 确定当前进程属于哪个路径集处理组
    path_set_index = rank // procs_per_group

    # 在组内的相对进程编号
    group_rank = rank % procs_per_group

    # 创建进程组
    group_ranks = list(range(path_set_index * procs_per_group, (path_set_index + 1) * procs_per_group))

    # 创建当前路径集处理子组的通信器
    group = comm.group.Incl(group_ranks)
    group_comm = comm.Create(group)

    # 开始计时
    start_time = time.time()

    # 获取当前进程组要处理的目标路径集
    global_target_paths = targetPathsSets[path_set_index]

    # --- 随机配对算法 ---
    # 只在全局进程0中进行配对计算
    if rank == 0:
        # 使用随机配对算法
        # 创建所有组的索引列表
        all_groups = list(range(num_path_sets))

        # 打乱顺序进行随机配对
        random.shuffle(all_groups)

        # 初始化配对字典
        p1_pairing = {}
        p2_pairing = {}
        p3_pairing = {}
        global_pairing = {}

        # 两两配对（进程1, 2, 3, 全局都使用相同的随机配对）
        for i in range(0, len(all_groups) - 1, 2):
            group_a = all_groups[i]
            group_b = all_groups[i + 1]

            # 进程1的配对
            p1_pairing[group_a] = group_b
            p1_pairing[group_b] = group_a

            # 进程2的配对
            p2_pairing[group_a] = group_b
            p2_pairing[group_b] = group_a

            # 进程3的配对
            p3_pairing[group_a] = group_b
            p3_pairing[group_b] = group_a

            # 全局PSO的配对
            global_pairing[group_a] = group_b
            global_pairing[group_b] = group_a

        # 如果组数是奇数，最后一个组没有配对
        if len(all_groups) % 2 == 1:
            last_group = all_groups[-1]
            p1_pairing[last_group] = None
            p2_pairing[last_group] = None
            p3_pairing[last_group] = None
            global_pairing[last_group] = None

        # 将配对信息发送给其他组的组内进程0
        for group_id in range(1, num_path_sets):  # 从1开始，因为0是当前进程
            group_master_rank = group_id * procs_per_group
            pairing_info = {
                'p1_partner': p1_pairing.get(group_id, None),
                'p2_partner': p2_pairing.get(group_id, None),
                'p3_partner': p3_pairing.get(group_id, None),
                'global_partner': global_pairing.get(group_id, None)
            }
            comm.send(pairing_info, dest=group_master_rank, tag=555)

        # 保存当前组（组0）的配对信息
        current_group_info = {
            'p1_partner': p1_pairing.get(0, None),
            'p2_partner': p2_pairing.get(0, None),
            'p3_partner': p3_pairing.get(0, None),
            'global_partner': global_pairing.get(0, None)
        }

    # 其他组的组内进程0接收配对信息
    if group_rank == 0 and rank != 0:
        status = MPI.Status()
        pairing_info = comm.recv(source=0, tag=555, status=status)
        current_group_info = pairing_info
    else:
        # 为确保在进程0也定义了current_group_info变量
        if rank != 0:
            current_group_info = None

    # 将配对信息广播给组内进程1、2和3
    if group_rank == 0:
        # 如果不是全局进程0，则需要先接收配对信息
        if rank != 0:
            p1_partner = current_group_info['p1_partner']
            p2_partner = current_group_info['p2_partner']
            p3_partner = current_group_info['p3_partner']
            global_partner = current_group_info['global_partner']
        else:
            p1_partner = current_group_info['p1_partner']
            p2_partner = current_group_info['p2_partner']
            p3_partner = current_group_info['p3_partner']
            global_partner = current_group_info['global_partner']

        # 计算随机配对伙伴实际rank并发送给组内进程
        if p1_partner is not None:
            p1_partner_rank = p1_partner * procs_per_group + 1  # 计算伙伴进程1的全局rank
            group_comm.send(p1_partner_rank, dest=1, tag=666)
        else:
            group_comm.send(None, dest=1, tag=666)

        if p2_partner is not None:
            p2_partner_rank = p2_partner * procs_per_group + 2  # 计算伙伴进程2的全局rank
            group_comm.send(p2_partner_rank, dest=2, tag=667)
        else:
            group_comm.send(None, dest=2, tag=667)

        # 发送进程3的随机配对伙伴信息
        if p3_partner is not None:
            p3_partner_rank = p3_partner * procs_per_group + 3  # 计算伙伴进程3的全局rank
            group_comm.send(p3_partner_rank, dest=3, tag=670)
        else:
            group_comm.send(None, dest=3, tag=670)

    # 计算全局PSO的伙伴rank
    global_partner_rank = None
    if group_rank == 0:
        # 随机配对伙伴
        if global_partner is not None:
            global_partner_rank = global_partner * procs_per_group  # 伙伴的进程0

    # 根据组内rank初始化相应的PSO
    if group_rank == 1:
        # 进程1 - 全路径生成
        local_target_path = global_target_paths[1]

        # 接收来自进程0的随机配对信息
        status = MPI.Status()
        partner_rank = group_comm.recv(source=0, tag=666, status=status)

        pso_local = PSO(
            mode="full",
            target_path=local_target_path,
            global_comm=comm,
            partner_rank=partner_rank,
            group_color=path_set_index,
            local_rank=1
        )

    elif group_rank == 2:
        # 进程2 - 部分路径生成
        local_target_path = global_target_paths[2]

        # 接收来自进程0的随机配对信息
        status = MPI.Status()
        partner_rank = group_comm.recv(source=0, tag=667, status=status)

        pso_local = PSO(
            mode="partial",
            target_path=local_target_path,
            global_comm=comm,
            partner_rank=partner_rank,
            group_color=path_set_index,
            local_rank=2
        )

    elif group_rank == 3:
        # 进程3 - path3生成
        local_target_path = global_target_paths[3]

        # 接收来自进程0的随机配对信息
        status = MPI.Status()
        partner_rank = group_comm.recv(source=0, tag=670, status=status)

        pso_local = PSO(
            mode="path3",
            target_path=local_target_path,
            global_comm=comm,
            partner_rank=partner_rank,
            group_color=path_set_index,
            local_rank=3
        )

    elif group_rank == 0:
        # 进程0 - 全局PSO (包含全局PSO之间的通信功能)
        pso_global = PSO5D(
            global_target_paths,
            group_comm,
            global_comm=comm,
            partner_rank=global_partner_rank,
            group_color=path_set_index
        )

    # 重置迭代计数和状态
    iteration = 1
    particle_sharing_count = 0
    solution_found = False

    # --- Main optimization loop with modified termination handling ---
    while iteration < MAX_ITERATIONS and not solution_found:
        if iteration % GENL == 0:  # 使用GENL控制局部优化周期
            if iteration > 0:  # 跳过第一次迭代，因为此时局部PSO还没有进行过优化
                if group_rank == 1:
                    top_particles = pso_local.getTopParticles(top_n=2)
                    top_data = [(p.position[0], p.position[1], p.position[2], p.fitness) for p in top_particles]
                    group_comm.send(top_data, dest=0, tag=101)

                elif group_rank == 2:
                    top_particles = pso_local.getTopParticles(top_n=2)
                    top_data = [(p.position[0], p.position[1], p.position[2], p.fitness) for p in top_particles]
                    group_comm.send(top_data, dest=0, tag=102)

                elif group_rank == 3:
                    top_particles = pso_local.getTopParticles(top_n=2)
                    top_data = [(p.position[0], p.position[1], p.position[2], p.fitness) for p in top_particles]
                    group_comm.send(top_data, dest=0, tag=103)

                elif group_rank == 0:
                    status = MPI.Status()
                    p1_particles = group_comm.recv(source=1, tag=101, status=status)
                    p2_particles = group_comm.recv(source=2, tag=102, status=status)
                    p3_particles = group_comm.recv(source=3, tag=103, status=status)

                    num_particles = pso_global.initializeSwarmFromLocalPSO(p1_particles, p2_particles, p3_particles)
                    particle_sharing_count += 1

            for global_step in range(GENO):  # 使用GENO控制全局优化迭代次数
                if group_rank == 0:
                    pso_global.updateParticles(global_target_paths)

                    if pso_global.gBestFitness == OPTIMAL_FITNESS:
                        solution_found = True
                        # 通知组内其他进程找到了最优解，需要终止
                        for i in range(1, procs_per_group):
                            group_comm.send(True, dest=i, tag=999)
                        # 通知配对的全局PSO伙伴找到了最优解
                        if pso_global.partner_rank is not None and not pso_global.partner_finished:
                            pso_global.notifyPartnerOfGlobalTermination()
                        break
                    else:
                        # 即使没有找到最优解，也要发送一个标志，保持进程同步
                        for i in range(1, procs_per_group):
                            group_comm.send(False, dest=i, tag=999)

                elif group_rank == 1:
                    for i in range(PSO5D.SWARM_SIZE):
                        status = MPI.Status()
                        first_val = group_comm.recv(source=0, tag=1, status=status)
                        second_val = group_comm.recv(source=0, tag=2, status=status)
                        third_val = group_comm.recv(source=0, tag=3, status=status)
                        path = mutation_process1(first_val, second_val, third_val)
                        group_comm.send(path, dest=0, tag=11)
                    # 接收是否找到最优解的通知
                    status = MPI.Status()
                    solution_found = group_comm.recv(source=0, tag=999, status=status)
                    # 更新局部PSO的全局解状态
                    pso_local.global_solution_found = solution_found
                    # 如果全局找到最优解，则通知伙伴进程
                    if solution_found:
                        pso_local.notifyPartnerOfGlobalTermination()

                elif group_rank == 2:
                    for i in range(PSO5D.SWARM_SIZE):
                        status = MPI.Status()
                        first_val = group_comm.recv(source=0, tag=1, status=status)
                        second_val = group_comm.recv(source=0, tag=2, status=status)
                        third_val = group_comm.recv(source=0, tag=3, status=status)
                        path = mutation_process2(first_val, second_val, third_val)
                        group_comm.send(path, dest=0, tag=21)
                    # 接收是否找到最优解的通知
                    status = MPI.Status()
                    solution_found = group_comm.recv(source=0, tag=999, status=status)
                    # 更新局部PSO的全局解状态
                    pso_local.global_solution_found = solution_found
                    # 如果全局找到最优解，则通知伙伴进程
                    if solution_found:
                        pso_local.notifyPartnerOfGlobalTermination()

                elif group_rank == 3:
                    for i in range(PSO5D.SWARM_SIZE):
                        status = MPI.Status()
                        first_val = group_comm.recv(source=0, tag=1, status=status)
                        second_val = group_comm.recv(source=0, tag=2, status=status)
                        third_val = group_comm.recv(source=0, tag=3, status=status)
                        path = mutation_process3(first_val, second_val, third_val)
                        group_comm.send(path, dest=0, tag=31)
                    # 接收是否找到最优解的通知
                    status = MPI.Status()
                    solution_found = group_comm.recv(source=0, tag=999, status=status)
                    # 更新局部PSO的全局解状态
                    pso_local.global_solution_found = solution_found
                    # 如果全局找到最优解，则通知伙伴进程
                    if solution_found:
                        pso_local.notifyPartnerOfGlobalTermination()

                # 如果找到了最优解，所有进程退出当前循环
                if solution_found:
                    break

                group_comm.Barrier()

                # 检查是否达到最大迭代
                if iteration + global_step >= MAX_ITERATIONS:
                    break

            # 全局优化结束后，从全局优化中选择H个适应度最高的粒子反馈给局部优化
            if group_rank == 0 and not solution_found:
                # 获取全局最优的H个粒子，分别提取为进程1、2和3需要的维度信息
                p1_feedback, p2_feedback, p3_feedback = pso_global.getTopParticlesForFeedback(top_n=H)

                # 发送反馈粒子给局部进程
                group_comm.send(p1_feedback, dest=1, tag=201)
                group_comm.send(p2_feedback, dest=2, tag=202)
                group_comm.send(p3_feedback, dest=3, tag=203)

            elif group_rank == 1 and not solution_found:
                # 组内进程1接收来自全局的反馈粒子
                status = MPI.Status()
                feedback_particles = group_comm.recv(source=0, tag=201, status=status)
                # 将反馈粒子整合到局部粒子群中
                pso_local.incorporateFeedbackParticles(feedback_particles)

            elif group_rank == 2 and not solution_found:
                # 组内进程2接收来自全局的反馈粒子
                status = MPI.Status()
                feedback_particles = group_comm.recv(source=0, tag=202, status=status)
                # 将反馈粒子整合到局部粒子群中
                pso_local.incorporateFeedbackParticles(feedback_particles)

            elif group_rank == 3 and not solution_found:
                # 组内进程3接收来自全局的反馈粒子
                status = MPI.Status()
                feedback_particles = group_comm.recv(source=0, tag=203, status=status)
                # 将反馈粒子整合到局部粒子群中
                pso_local.incorporateFeedbackParticles(feedback_particles)

            # 如果已经找到了最优解，则终止迭代
            if solution_found:
                break

            # 将迭代计数更新一次，表示完成了一个完整的全局优化阶段
            iteration += 1
            group_comm.Barrier()

        else:
            # 局部阶段：所有局部进程执行三维PSO更新
            if group_rank in [1, 2, 3]:
                pso_local.updateParticles()

            # 局部优化阶段每次迭代计数加1
            iteration += 1
            group_comm.Barrier()

        # 检查是否达到最大迭代次数
        if iteration >= MAX_ITERATIONS:
            # 即使达到最大迭代次数，也通知伙伴进程自己已经完成
            if group_rank in [1, 2, 3] and pso_local.partner_rank is not None and not pso_local.partner_finished:
                pso_local.notifyPartnerOfGlobalTermination()
            # 全局PSO也通知其伙伴
            if group_rank == 0 and pso_global.partner_rank is not None and not pso_global.partner_finished:
                pso_global.notifyPartnerOfGlobalTermination()
            break

    # 结束计时并计算耗时
    end_time = time.time()
    search_time = end_time - start_time

    # 收集实验结果
    if group_rank == 0:
        # 🔧 新增:收集所有局部PSO的评价次数
        eval_p1 = group_comm.recv(source=1, tag=301)
        eval_p2 = group_comm.recv(source=2, tag=302)
        eval_p3 = group_comm.recv(source=3, tag=303)

        # 🔧 新增:计算总评价次数
        total_evaluations = pso_global.evaluation_count + eval_p1 + eval_p2 + eval_p3

        # 只有组内进程0返回结果
        result = {
            'group_id': path_set_index,
            'success': solution_found,
            'time': search_time,  # ← 修改：无论成功失败都记录实际时间
            'iterations': iteration,  # 🔧 无论成功与否都记录迭代次数
            'evaluations': total_evaluations,  # 🔧 新增:评价次数
            'best_fitness': pso_global.gBestFitness,
            'best_position': pso_global.gBest.position if pso_global.gBest.position else None,
            'comm_time': pso_global.comm_time
        }
        return result
    elif group_rank in [1, 2, 3]:
        # 🔧 新增:局部进程发送评价次数给进程0
        if group_rank == 1:
            group_comm.send(pso_local.evaluation_count, dest=0, tag=301)
        elif group_rank == 2:
            group_comm.send(pso_local.evaluation_count, dest=0, tag=302)
        elif group_rank == 3:
            group_comm.send(pso_local.evaluation_count, dest=0, tag=303)
    else:
        return None


def main():
    comm = MPI.COMM_WORLD
    rank = comm.Get_rank()
    size = comm.Get_size()

    # 每个路径集需要的进程数
    procs_per_group = 4

    # 计算实际的路径集数量
    num_path_sets = len(targetPathsSets)
    max_processable_sets = size // procs_per_group

    if max_processable_sets < num_path_sets:
        if rank == 0:
            print(f"警告：定义了{num_path_sets}个路径集，但当前进程数{size}只能处理{max_processable_sets}个路径集。")
            print(f"只处理前{max_processable_sets}个路径集。")
        num_path_sets = max_processable_sets

    # 检查进程数是否足够
    if size < procs_per_group:
        if rank == 0:
            print(f"错误：至少需要{procs_per_group}个进程来处理一个路径集！")
        return

    # 开始总计时
    total_start_time = time.time()

    # 🔧 修改：初始化统计数据结构，添加适应值列表和详细记录列表，包含迭代次数
    if rank == 0:
        group_stats = {}
        for group_id in range(num_path_sets):
            group_stats[group_id] = {
                'successes': 0,
                'success_times': [],
                'success_iterations': [],
                'success_evaluations': [],  # 🔧 新增:成功实验的评价次数
                'total_comm_time': 0.0,
                'all_fitness_values': [],
                # 🔧 修改：用于Excel输出的详细记录，包含迭代次数
                'success_list': [],  # 是否成功（0/1）
                'time_list': [],  # 迭代时间
                'fitness_list': [],  # 最终适应度值
                'iterations_list': [],  # 迭代次数 ← 新增
                'evaluations_list': []  # 🔧 新增:评价次数
            }

    # 多次实验循环
    for experiment_id in range(NUM_EXPERIMENTS):
        if rank == 0:
            # 区分预热和正式实验的输出
            if experiment_id < WARMUP_EXPERIMENTS:
                print(f"\n=== 预热实验 {experiment_id + 1}/{WARMUP_EXPERIMENTS} (不计入统计) ===")
            else:
                formal_exp_id = experiment_id - WARMUP_EXPERIMENTS + 1
                formal_total = NUM_EXPERIMENTS - WARMUP_EXPERIMENTS
                print(f"\n=== 正式实验 {formal_exp_id}/{formal_total} (第{experiment_id + 1}次总实验) ===")

        # 运行单次实验
        experiment_result = run_single_experiment(
            experiment_id, comm, rank, size, num_path_sets, procs_per_group
        )

        # 收集实验结果
        if rank == 0:
            # 收集所有组的结果
            experiment_results = {}

            # 收集自己组（组0）的结果
            if experiment_result is not None:
                experiment_results[0] = experiment_result

            # 收集其他组的结果
            for group_id in range(1, num_path_sets):
                group_master_rank = group_id * procs_per_group
                try:
                    other_result = comm.recv(source=group_master_rank, tag=888)
                    experiment_results[group_id] = other_result
                except:
                    # 如果某个组没有结果，跳过
                    pass

            # 输出当前实验进度
            if experiment_id < WARMUP_EXPERIMENTS:
                progress_line = f"预热实验 {experiment_id + 1}: "
            else:
                formal_exp_id = experiment_id - WARMUP_EXPERIMENTS + 1
                progress_line = f"正式实验 {formal_exp_id}: "

            for group_id in range(num_path_sets):
                if group_id in experiment_results:
                    result = experiment_results[group_id]
                    if result['success']:
                        progress_line += f"组{group_id}[成功,{result['time']:.1f}s,{result['iterations']}次,评价={result['evaluations']},适应值={result['best_fitness']:.4f}] "
                    else:
                        progress_line += f"组{group_id}[失败,{result['time']:.1f}s,{result['iterations']}次,评价={result['evaluations']},适应值={result['best_fitness']:.4f}] "
                else:
                    progress_line += f"组{group_id}[无结果] "
            print(progress_line)

            # 只有正式实验才更新统计数据
            if experiment_id >= WARMUP_EXPERIMENTS:
                for group_id, result in experiment_results.items():
                    # 记录所有实验的适应值（包括成功和失败的）
                    group_stats[group_id]['all_fitness_values'].append(result['best_fitness'])

                    # 🔧 修改：记录每次实验的详细结果用于Excel输出，包含迭代次数 - 失败时记录0
                    # 修正：记录每次实验的详细结果用于Excel输出 - 无论成功失败都记录真实时间和迭代次数
                    group_stats[group_id]['success_list'].append(1 if result['success'] else 0)
                    group_stats[group_id]['time_list'].append(result['time'])  # ← 修改：直接记录实际时间
                    group_stats[group_id]['fitness_list'].append(result['best_fitness'])
                    group_stats[group_id]['iterations_list'].append(result['iterations'])  # ← 修改：直接记录实际迭代次数
                    group_stats[group_id]['evaluations_list'].append(result['evaluations'])  # 🔧 新增

                    if result['success']:
                        group_stats[group_id]['successes'] += 1
                        group_stats[group_id]['success_times'].append(result['time'])
                        group_stats[group_id]['success_iterations'].append(result['iterations'])
                        group_stats[group_id]['success_evaluations'].append(result['evaluations'])  # 🔧 新增
                        group_stats[group_id]['total_comm_time'] += result['comm_time']

        elif experiment_result is not None:
            # 其他组的组内进程0发送结果给全局进程0
            comm.send(experiment_result, dest=0, tag=888)

        # 同步所有进程
        comm.Barrier()

    # 结束总计时
    total_end_time = time.time()
    total_time = total_end_time - total_start_time

    # 输出最终统计结果
    if rank == 0:
        print(f"\n{'=' * 60}")
        print(f"=== 实验统计结果 (基于正式实验) ===")
        print(f"{'=' * 60}")

        # 计算正式实验次数
        formal_experiments = NUM_EXPERIMENTS - WARMUP_EXPERIMENTS
        print(f"预热实验: {WARMUP_EXPERIMENTS}次 (不计入统计)")
        print(f"正式实验: {formal_experiments}次 (用于统计)")
        print(f"总实验次数: {NUM_EXPERIMENTS}次")
        print("-" * 60)

        for group_id in range(num_path_sets):
            stats = group_stats[group_id]
            success_count = stats['successes']
            # 使用正式实验次数计算成功率
            success_rate = success_count / formal_experiments * 100 if formal_experiments > 0 else 0

            # 计算适应值统计（包括标准差）
            fitness_values = stats['all_fitness_values']
            if fitness_values:
                avg_fitness = sum(fitness_values) / len(fitness_values)
                std_dev_fitness = calculate_std_dev(fitness_values)
                max_fitness = max(fitness_values)
                min_fitness = min(fitness_values)
            else:
                avg_fitness = 0.0
                std_dev_fitness = 0.0
                max_fitness = 0.0
                min_fitness = 0.0

            if success_count > 0:
                avg_time = sum(stats['success_times']) / success_count
                avg_iterations = sum(stats['success_iterations']) / success_count
                avg_evaluations = sum(stats['success_evaluations']) / success_count  # 🔧 新增
                avg_comm_time = stats['total_comm_time'] / success_count
                print(f"组{group_id}: 成功率 {success_count}/{formal_experiments} ({success_rate:.1f}%), "
                      f"平均时间 {avg_time:.3f}秒, 平均迭代 {avg_iterations:.0f}次, "
                      f"平均评价 {avg_evaluations:.0f}次, "  # 🔧 新增
                      f"平均通信时间 {avg_comm_time:.3f}秒")
                print(f"       适应值统计 - 平均: {avg_fitness:.4f}, 标准差: {std_dev_fitness:.4f}, "
                      f"最大: {max_fitness:.4f}, 最小: {min_fitness:.4f}")
            else:
                print(f"组{group_id}: 成功率 {success_count}/{formal_experiments} ({success_rate:.1f}%), "
                      f"无成功实验")
                print(f"       适应值统计 - 平均: {avg_fitness:.4f}, 标准差: {std_dev_fitness:.4f}, "
                      f"最大: {max_fitness:.4f}, 最小: {min_fitness:.4f}")

        print(f"\n总运行时间：{total_time:.2f} 秒")
        print(f"处理了 {num_path_sets} 个路径集，使用了 {num_path_sets * procs_per_group} 个进程")

        # 输出整体统计（基于正式实验）
        total_successes = sum(stats['successes'] for stats in group_stats.values())
        total_formal_experiments = formal_experiments * num_path_sets
        overall_success_rate = total_successes / total_formal_experiments * 100 if total_formal_experiments > 0 else 0

        # 计算整体适应值统计
        all_fitness_values = []
        for stats in group_stats.values():
            all_fitness_values.extend(stats['all_fitness_values'])

        if all_fitness_values:
            overall_avg_fitness = sum(all_fitness_values) / len(all_fitness_values)
            overall_std_dev = calculate_std_dev(all_fitness_values)
            overall_max_fitness = max(all_fitness_values)
            overall_min_fitness = min(all_fitness_values)
        else:
            overall_avg_fitness = 0.0
            overall_std_dev = 0.0
            overall_max_fitness = 0.0
            overall_min_fitness = 0.0

        # 🔧 新增:计算整体评价次数统计
        all_evaluations = []
        for stats in group_stats.values():
            all_evaluations.extend(stats['evaluations_list'])

        if all_evaluations:
            overall_avg_evaluations = sum(all_evaluations) / len(all_evaluations)
            overall_std_dev_evaluations = calculate_std_dev(all_evaluations)
            overall_max_evaluations = max(all_evaluations)
            overall_min_evaluations = min(all_evaluations)
        else:
            overall_avg_evaluations = 0.0
            overall_std_dev_evaluations = 0.0
            overall_max_evaluations = 0
            overall_min_evaluations = 0

        print(f"整体成功率：{total_successes}/{total_formal_experiments} ({overall_success_rate:.1f}%)")
        print(f"整体适应值统计 - 平均: {overall_avg_fitness:.4f}, 标准差: {overall_std_dev:.4f}, "
              f"最大: {overall_max_fitness:.4f}, 最小: {overall_min_fitness:.4f}")
        # 🔧 新增:输出整体评价次数统计
        print(f"整体评价次数统计 - 平均: {overall_avg_evaluations:.0f}, 标准差: {overall_std_dev_evaluations:.0f}, "
              f"最大: {overall_max_evaluations}, 最小: {overall_min_evaluations}")
        print(f"(基于 {formal_experiments} 次正式实验，排除 {WARMUP_EXPERIMENTS} 次预热实验)")

        # 🔧 修改：保存详细结果到Excel文件，包含迭代次数
        print(f"\n{'=' * 60}")
        print("=== 保存实验结果到Excel文件 ===")
        print(f"{'=' * 60}")
        save_all_results_to_excel(group_stats, num_path_sets)
        print("Excel文件保存完成！")


if __name__ == "__main__":
    main()